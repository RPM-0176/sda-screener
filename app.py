ADMIN_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Admin - SDA Screener</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Arial,sans-serif;background:#f3f4f6;color:#111}
.top{background:#002060;color:#fff;padding:14px 24px;display:flex;justify-content:space-between;align-items:center}
.top h1{font-size:15px;font-weight:700}
.top a{color:#93c5fd;font-size:13px;text-decoration:none;margin-left:16px}
.wrap{max-width:1200px;margin:0 auto;padding:20px}
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px}
.stat{background:#fff;border-radius:8px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.1)}
.stat .n{font-size:28px;font-weight:800;color:#002060}
.stat .l{font-size:11px;color:#6B7280;margin-top:4px;text-transform:uppercase;font-weight:600}
.tabs{display:flex;gap:4px;margin-bottom:16px}
.tab{padding:10px 18px;border-radius:6px;font-size:13px;font-weight:600;background:#e5e7eb;color:#374151;text-decoration:none;display:inline-block}
.tab.on{background:#002060;color:#fff}
.card{background:#fff;border-radius:8px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.1);margin-bottom:16px}
.card h2{font-size:13px;font-weight:700;color:#002060;margin-bottom:12px;padding-bottom:8px;border-bottom:2px solid #e5e7eb}
table{width:100%;border-collapse:collapse;font-size:12px}
th{text-align:left;padding:8px 10px;background:#f9fafb;color:#6B7280;font-size:11px;font-weight:700;text-transform:uppercase;border-bottom:1px solid #e5e7eb}
td{padding:8px 10px;border-bottom:1px solid #f3f4f6;vertical-align:middle}
.badge{display:inline-block;padding:2px 6px;border-radius:10px;font-size:10px;font-weight:700}
.ba{background:#fef3c7;color:#92400e}
.bt{background:#e0f2fe;color:#0369a1}
.bact{background:#dcfce7;color:#166534}
.bina{background:#fee2e2;color:#991b1b}
.btn{padding:5px 10px;border:none;border-radius:5px;cursor:pointer;font-size:11px;font-weight:700;text-decoration:none;display:inline-block}
.blue{background:#002060;color:#fff}
.red{background:#fee2e2;color:#991b1b}
.grn{background:#dcfce7;color:#166534}
form.inline{display:inline}
.addform{background:#fff;border-radius:8px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.1);margin-bottom:16px}
.addform .grid{display:grid;grid-template-columns:1fr 1fr 1fr 1fr auto;gap:10px;align-items:end}
.addform label{display:block;font-size:11px;font-weight:700;color:#6B7280;margin-bottom:4px;text-transform:uppercase}
.addform input,.addform select{width:100%;padding:8px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:13px}
.pwfield{padding:6px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:12px;width:140px}
.msg{padding:10px 14px;border-radius:6px;margin-bottom:14px;font-size:13px}
.ok{background:#dcfce7;color:#166534;border:1px solid #86efac}
.er{background:#fee2e2;color:#991b1b;border:1px solid #fca5a5}
</style>
</head>
<body>
<div class="top">
  <h1>SDA Screener - Admin Panel</h1>
  <div>
    <span style="font-size:13px">{{ current_user.full_name }}</span>
    <a href="/dashboard">Dashboard</a>
    <a href="/logout">Sign out</a>
  </div>
</div>
<div class="wrap">
  {% if msg %}<div class="msg {{ msg_type }}">{{ msg }}</div>{% endif %}
  <div class="stats">
    <div class="stat"><div class="n">{{ stats.total_users }}</div><div class="l">Team members</div></div>
    <div class="stat"><div class="n">{{ stats.total_sessions }}</div><div class="l">Sessions</div></div>
    <div class="stat"><div class="n">{{ stats.views }}</div><div class="l">Property views</div></div>
    <div class="stat"><div class="n">{{ stats.downloads }}</div><div class="l">Downloads</div></div>
  </div>
  <div class="tabs">
    <a href="/admin/activity" class="tab {% if section=='activity' %}on{% endif %}">Team activity</a>
    <a href="/admin/props" class="tab {% if section=='props' %}on{% endif %}">Property views</a>
    <a href="/admin/downloads" class="tab {% if section=='downloads' %}on{% endif %}">Downloads</a>
    <a href="/admin/hidden" class="tab {% if section=='hidden' %}on{% endif %}">Hidden properties{% if hidden_count %} ({{ hidden_count }}){% endif %}</a>
    <a href="/admin/users" class="tab {% if section=='users' %}on{% endif %}">Manage users</a>
    <a href="/admin/upload_page" class="tab" style="background:#0F6E56;color:#fff">Upload CSV data</a>
    <a href="/admin/upload_sda_page" class="tab" style="background:#185FA5;color:#fff">Upload SDA market</a>
    <a href="/admin/upload_supply_page" class="tab" style="background:#0F6E56;color:#fff">Upload SDA supply</a>
  </div>

  {% if section == 'activity' %}
  <div class="card">
    <h2>Team member activity</h2>
    <table><thead><tr><th>Name</th><th>Username</th><th>Role</th><th>Sessions</th><th>Last login</th><th>Status</th><th>Change password</th></tr></thead>
    <tbody>
    {% for u in users %}
    <tr>
      <td><strong>{{ u.full_name }}</strong></td>
      <td>{{ u.username }}</td>
      <td><span class="badge {{ 'ba' if u.role=='admin' else 'bt' }}">{{ u.role }}</span></td>
      <td>{{ u.session_count or 0 }}</td>
      <td>{{ u.last_login[:16] if u.last_login else 'Never' }}</td>
      <td><span class="badge {{ 'bact' if u.active else 'bina' }}">{{ 'Active' if u.active else 'Disabled' }}</span></td>
      <td>
        <form class="inline" method="POST" action="/admin/users/{{ u.id }}/password">
          <input type="password" name="password" placeholder="New password" class="pwfield">
          <button type="submit" class="btn blue">Update</button>
        </form>
      </td>
    </tr>
    {% else %}
    <tr><td colspan="7" style="color:#9CA3AF;padding:12px">No team members yet. Go to Manage Users to add them.</td></tr>
    {% endfor %}
    </tbody></table>
  </div>
  {% endif %}

  {% if section == 'props' %}
  <div class="card">
    <h2>Property views</h2>
    <table><thead><tr><th>When</th><th>User</th><th>Address</th><th>State</th><th>SA4</th><th>Price</th></tr></thead>
    <tbody>
    {% for e in prop_views %}
    <tr><td>{{ e.created_at[:16] }}</td><td>{{ e.full_name }}</td><td>{{ e.addr }}</td><td>{{ e.state }}</td><td>{{ e.sa4 }}</td><td>{{ e.price }}</td></tr>
    {% else %}
    <tr><td colspan="6" style="color:#9CA3AF;padding:12px">No property views logged yet</td></tr>
    {% endfor %}
    </tbody></table>
  </div>
  {% endif %}

  {% if section == 'downloads' %}
  <div class="card">
    <h2>Downloads</h2>
    <table><thead><tr><th>When</th><th>User</th><th>Type</th><th>Address</th><th>State</th></tr></thead>
    <tbody>
    {% for e in downloads %}
    <tr><td>{{ e.created_at[:16] }}</td><td>{{ e.full_name }}</td><td>{{ e.event_type }}</td><td>{{ e.addr }}</td><td>{{ e.state }}</td></tr>
    {% else %}
    <tr><td colspan="5" style="color:#9CA3AF;padding:12px">No downloads logged yet</td></tr>
    {% endfor %}
    </tbody></table>
  </div>
  {% endif %}

  {% if section == 'hidden' %}
  <div class="card">
    <h2>Hidden properties</h2>
    <p style="font-size:11px;color:#6B7280;margin-bottom:10px">
      Team members can hide a property from the screener by giving a reason. Restore puts it back in the screener. Hard delete removes it from the CSV permanently &mdash; this cannot be undone.
    </p>
    <table>
      <thead><tr>
        <th>When</th><th>Hidden by</th><th>State</th><th>Address</th><th>Reason</th><th style="text-align:right">Action</th>
      </tr></thead>
      <tbody>
      {% for h in hidden %}
        <tr id="hide-row-{{ h.id }}">
          <td>{{ h.hidden_at[:16] }}</td>
          <td>{{ h.hidden_by_name }}</td>
          <td><span class="badge bt">{{ h.state|upper }}</span></td>
          <td>{{ h.address }}</td>
          <td style="max-width:340px;white-space:normal">{{ h.reason }}</td>
          <td style="text-align:right;white-space:nowrap">
            <button type="button" class="btn grn" onclick="hideAction({{ h.id }}, 'restore')">Restore</button>
            <button type="button" class="btn red" data-state="{{ h.state|e }}" data-addr="{{ h.address|e }}" onclick="hideAction({{ h.id }}, 'harddelete', this)">Hard delete</button>
          </td>
        </tr>
      {% else %}
        <tr><td colspan="6" style="color:#9CA3AF;padding:12px">No properties are currently hidden.</td></tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
  <script>
  function adminToast(text){
    var t = document.getElementById('admin-toast');
    if(!t){
      t = document.createElement('div');
      t.id = 'admin-toast';
      t.style.cssText = 'position:fixed;right:20px;bottom:20px;z-index:9999;background:#0F6E56;color:#fff;padding:12px 16px;border-radius:8px;font-weight:600;font-size:14px;box-shadow:0 2px 12px rgba(0,0,0,0.2);max-width:380px;line-height:1.35';
      document.body.appendChild(t);
    }
    t.textContent = text;
    t.style.display = 'block';
    clearTimeout(window._adminToastTimer);
    window._adminToastTimer = setTimeout(function(){ t.style.display='none'; }, 5000);
  }
  function hideAction(id, kind, btn){
    if(kind === 'restore'){
      if(!confirm('Restore this property to the screener?')) return;
      fetch('/api/unhide-property', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({id: id})
      }).then(function(r){return r.json().then(function(j){return {ok:r.ok, j:j};});})
        .then(function(res){
          if(res.ok){
            var row = document.getElementById('hide-row-'+id);
            if(row) row.remove();
            adminToast('Restored to the screener. Reload your dashboard tab to see it (clear any filters if it does not appear).');
          } else {
            alert('Restore failed: '+(res.j && res.j.error || 'unknown'));
          }
        }).catch(function(e){ alert('Restore error: '+e); });
    } else if(kind === 'harddelete'){
      var state = (btn && btn.getAttribute('data-state')) || '';
      var address = (btn && btn.getAttribute('data-addr')) || '';
      if(!confirm('PERMANENTLY delete this property from the '+state.toUpperCase()+' CSV: '+address+'. This cannot be undone; re-uploading the CSV will bring it back.')) return;
      fetch('/api/hard-delete-property', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({state: state, address: address})
      }).then(function(r){return r.json().then(function(j){return {ok:r.ok, j:j};});})
        .then(function(res){
          if(res.ok){
            var row = document.getElementById('hide-row-'+id);
            if(row) row.remove();
            adminToast('Permanently removed from the '+state.toUpperCase()+' CSV.');
          } else {
            alert('Hard delete failed: '+(res.j && res.j.error || 'unknown'));
          }
        }).catch(function(e){ alert('Hard delete error: '+e); });
    }
  }
  </script>
  {% endif %}

  {% if section == 'users' %}
  <div class="addform">
    <h2 style="margin-bottom:16px">Add team member</h2>
    <form method="POST" action="/admin/users/add_form">
      <div class="grid">
        <div><label>Full name</label><input name="full_name" placeholder="e.g. Sarah Johnson" required></div>
        <div><label>Username</label><input name="username" placeholder="e.g. sarah" required></div>
        <div><label>Password</label><input type="password" name="password" placeholder="Min 6 characters" required></div>
        <div><label>Role</label><select name="role"><option value="team">Team member</option><option value="admin">Admin</option></select></div>
        <div><label>&nbsp;</label><button type="submit" class="btn blue" style="width:100%;padding:9px">Add</button></div>
      </div>
    </form>
  </div>
  <div class="card">
    <h2>All users</h2>
    <table><thead><tr><th>Name</th><th>Username</th><th>Role</th><th>Status</th><th>Change password</th><th>Access</th></tr></thead>
    <tbody>
    {% for u in all_users %}
    <tr>
      <td><strong>{{ u.full_name }}</strong></td>
      <td>{{ u.username }}</td>
      <td>
        <span class="badge {{ 'ba' if u.role=='admin' else 'bt' }}">{{ u.role }}</span>
        {% if u.id != current_user.user_id %}
        <form class="inline" method="POST" action="/admin/users/{{ u.id }}/toggle_role" style="display:inline-block;margin-left:6px"
              onsubmit="return confirm('Change {{ u.full_name }} to {{ 'team member' if u.role=='admin' else 'admin' }}?')">
          <button type="submit" class="btn" style="padding:3px 9px;font-size:11px;background:#fff;border:1px solid #185FA5;color:#185FA5;font-weight:600">{{ '\u2192 Team' if u.role=='admin' else '\u2192 Admin' }}</button>
        </form>
        {% endif %}
      </td>
      <td><span class="badge {{ 'bact' if u.active else 'bina' }}">{{ 'Active' if u.active else 'Disabled' }}</span></td>
      <td>
        <form class="inline" method="POST" action="/admin/users/{{ u.id }}/password">
          <input type="password" name="password" placeholder="New password" class="pwfield">
          <button type="submit" class="btn blue">Update</button>
        </form>
      </td>
      <td>
        <form class="inline" method="POST" action="/admin/users/{{ u.id }}/toggle">
          <button type="submit" class="btn {{ 'red' if u.active else 'grn' }}">{{ 'Disable' if u.active else 'Enable' }}</button>
        </form>
      </td>
    </tr>
    {% else %}
    <tr><td colspan="6" style="color:#9CA3AF;padding:12px">No users yet</td></tr>
    {% endfor %}
    </tbody></table>
  </div>
  {% endif %}

</div>
</body>
</html>"""

UPLOAD_HTML = """<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Upload Data - SDA Screener</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Arial,sans-serif;background:#f3f4f6}
.top{background:#002060;color:#fff;padding:14px 24px;display:flex;justify-content:space-between;align-items:center}
.top h1{font-size:15px;font-weight:700}
.top a{color:#93c5fd;font-size:13px;text-decoration:none;margin-left:16px}
.wrap{max-width:900px;margin:0 auto;padding:24px}
.card{background:#fff;border-radius:8px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.1);margin-bottom:16px}
.card h2{font-size:14px;font-weight:700;color:#002060;margin-bottom:16px;padding-bottom:8px;border-bottom:2px solid #e5e7eb}
.upload-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:20px}
.state-card{border:2px solid #e5e7eb;border-radius:8px;padding:16px;text-align:center}
.state-card.loaded{border-color:#0F6E56;background:#f0fdf4}
.state-name{font-size:18px;font-weight:800;color:#002060;margin-bottom:8px}
.state-rows{font-size:13px;color:#0F6E56;font-weight:600;margin-bottom:4px}
.state-date{font-size:11px;color:#6B7280;margin-bottom:12px}
.state-none{font-size:12px;color:#9CA3AF;margin-bottom:12px}
.upload-form{display:flex;flex-direction:column;gap:8px}
.upload-form input[type=file]{font-size:12px}
.btn{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-size:12px;font-weight:700}
.blue{background:#002060;color:#fff;width:100%}
.msg{padding:10px 14px;border-radius:6px;margin-bottom:16px;font-size:13px}
.ok{background:#dcfce7;color:#166534;border:1px solid #86efac}
.er{background:#fee2e2;color:#991b1b;border:1px solid #fca5a5}
</style>
</head>
<body>
<div class="top">
  <h1>SDA Screener - Upload Property Data</h1>
  <div>
    <a href="/admin">Admin panel</a>
    <a href="/dashboard">Dashboard</a>
    <a href="/logout">Sign out</a>
  </div>
</div>
<div class="wrap">
  <div id="msg-area"></div>
  <div class="card">
    <h2>Upload CSV files for each state</h2>
    <p style="font-size:12px;color:#6B7280;margin-bottom:16px">Upload the *_screener_ready.csv files generated by process_rp_data.py. Once uploaded, all team members will see the data automatically when they log in.</p>
    <div class="upload-grid" id="state-grid"></div>
  </div>
</div>
<script>
var status = {{STATUS}};
var msg = '{{MSG}}';
var msgType = '{{MSG_TYPE}}';
var states = ['vic','nsw','qld'];
var labels = {vic:'VIC',nsw:'NSW',qld:'QLD'};

if(msg){
  var d = document.getElementById('msg-area');
  d.innerHTML = '<div class="msg '+msgType+'">'+decodeURIComponent(msg.replace(/[+]/g,' '))+'</div>';
}

var grid = document.getElementById('state-grid');
states.forEach(function(st){
  var s = status[st];
  var loaded = s && s.rows > 0;
  var html = '<div class="state-card '+(loaded?'loaded':'')+'">'
    +'<div class="state-name">'+labels[st]+'</div>';
  if(loaded){
    html += '<div class="state-rows">'+s.rows+' properties</div>'
      +'<div class="state-date">Uploaded: '+s.uploaded_at.substring(0,16)+' by '+s.uploaded_by+'</div>';
  } else {
    html += '<div class="state-none">No data uploaded yet</div>';
  }
  html += '<form class="upload-form" method="POST" action="/admin/upload" enctype="multipart/form-data">'
    +'<input type="hidden" name="state" value="'+st+'">'
    +'<input type="file" name="csvfile" accept=".csv" required>'
    +'<button type="submit" class="btn blue">'+(loaded?'Replace ':'Upload ')+labels[st]+' CSV</button>'
    +'</form></div>';
  grid.innerHTML += html;
});
</script>
</body></html>"""

UPLOAD_SDA_HTML = """<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Upload SDA Market - SDA Screener</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Arial,sans-serif;background:#f3f4f6}
.top{background:#002060;color:#fff;padding:14px 24px;display:flex;justify-content:space-between;align-items:center}
.top h1{font-size:15px;font-weight:700}
.top a{color:#93c5fd;font-size:13px;text-decoration:none;margin-left:16px}
.wrap{max-width:1100px;margin:0 auto;padding:24px}
.card{background:#fff;border-radius:8px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.1);margin-bottom:16px}
.card h2{font-size:14px;font-weight:700;color:#002060;margin-bottom:8px;padding-bottom:8px;border-bottom:2px solid #e5e7eb}
.layer-section{margin-bottom:24px}
.layer-title{font-size:13px;font-weight:700;color:#002060;margin:18px 0 10px;padding:6px 10px;border-radius:4px;display:inline-flex;align-items:center;gap:6px}
.layer-title.radius{background:#dbeafe;color:#1e40af}
.layer-title.existing{background:#dcfce7;color:#166534}
.layer-title.ghomes{background:#ede9fe;color:#5b21b6}
.layer-title.permitted{background:#fef9c3;color:#854d0e}
.dot{width:10px;height:10px;border-radius:50%}
.dot.radius{background:#1e40af}
.dot.existing{background:#0F6E56}
.dot.ghomes{background:#7C3AED}
.dot.permitted{background:#EAB308}
.upload-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:8px}
.state-card{border:2px solid #e5e7eb;border-radius:8px;padding:12px;text-align:center}
.state-card.loaded{border-color:#0F6E56;background:#f0fdf4}
.state-name{font-size:14px;font-weight:800;color:#002060;margin-bottom:6px}
.state-pins{font-size:12px;color:#0F6E56;font-weight:600;margin-bottom:2px}
.state-date{font-size:10px;color:#6B7280;margin-bottom:8px}
.state-none{font-size:11px;color:#9CA3AF;margin-bottom:8px}
.upload-form{display:flex;flex-direction:column;gap:6px}
.upload-form input[type=file]{font-size:11px}
.btn{padding:7px 12px;border:none;border-radius:6px;cursor:pointer;font-size:11px;font-weight:700}
.blue{background:#002060;color:#fff;width:100%}
.msg{padding:10px 14px;border-radius:6px;margin-bottom:16px;font-size:13px}
.ok{background:#dcfce7;color:#166534;border:1px solid #86efac}
.er{background:#fee2e2;color:#991b1b;border:1px solid #fca5a5}
.help{font-size:11px;color:#6B7280;line-height:1.5;background:#f9fafb;padding:10px 14px;border-radius:6px;margin-bottom:14px;border-left:3px solid #185FA5}
</style>
</head>
<body>
<div class="top">
  <h1>SDA Screener - Upload SDA Market Data</h1>
  <div>
    <a href="/admin">Admin panel</a>
    <a href="/admin/upload_page">Property data</a>
    <a href="/dashboard">Dashboard</a>
    <a href="/logout">Sign out</a>
  </div>
</div>
<div class="wrap">
  <div id="msg-area"></div>

  <div class="card">
    <h2>SDA Market Map data</h2>
    <p class="help">Four pin layers for the SDA Market Map: <strong>Radius</strong> (your own pipeline + built projects), <strong>Existing SDA</strong> (operating, filled SDA dwellings), <strong>Group Homes</strong> (non-SDA disability accommodation), <strong>Permitted/approved</strong> (approved but not yet built). Upload one CSV per layer-state combination. Each upload <strong>replaces</strong> all pins for that layer+state. CSVs should be LandChecker-format with at minimum: Address, Suburb, State, Coordinates (the lng,lat pair).</p>
    <div id="layer-areas"></div>
  </div>
</div>
<script>
var status = {{STATUS}};
var msg = '{{MSG}}';
var msgType = '{{MSG_TYPE}}';
var states = ['vic','nsw','qld'];
var stateLabels = {vic:'VIC',nsw:'NSW',qld:'QLD'};
var layers = [
  {key:'radius',    label:'Radius pipeline / built'},
  {key:'existing',  label:'Existing SDA'},
  {key:'ghomes',    label:'Group Homes'},
  {key:'permitted', label:'Permitted / approved'}
];
if(msg){
  var d = document.getElementById('msg-area');
  d.innerHTML = '<div class="msg '+msgType+'">'+decodeURIComponent(msg.replace(/[+]/g,' '))+'</div>';
}

var areas = document.getElementById('layer-areas');
layers.forEach(function(L){
  var section = '<div class="layer-section">'
    + '<div class="layer-title '+L.key+'"><span class="dot '+L.key+'"></span>'+L.label+'</div>'
    + '<div class="upload-grid">';
  states.forEach(function(st){
    var s = (status[L.key] && status[L.key][st]) || null;
    var loaded = s && s.pins > 0;
    var html = '<div class="state-card '+(loaded?'loaded':'')+'">'
      + '<div class="state-name">'+stateLabels[st]+'</div>';
    if(loaded){
      html += '<div class="state-pins">'+s.pins+' pins</div>'
        + '<div class="state-date">Uploaded: '+s.uploaded_at.substring(0,16)+' by '+s.uploaded_by+'</div>';
    } else {
      html += '<div class="state-none">No data uploaded</div>';
    }
    html += '<form class="upload-form" method="POST" action="/admin/upload_sda" enctype="multipart/form-data">'
      + '<input type="hidden" name="layer" value="'+L.key+'">'
      + '<input type="hidden" name="state" value="'+st+'">'
      + '<input type="file" name="csvfile" accept=".csv" required>'
      + '<button type="submit" class="btn blue">'+(loaded?'Replace ':'Upload ')+stateLabels[st]+'</button>'
      + '</form></div>';
    section += html;
  });
  section += '</div></div>';
  areas.innerHTML += section;
});
</script>
</body></html>"""

"""SDA Property Screener - Team Server"""
from flask import Flask, request, session, redirect, url_for, jsonify, send_file, abort, make_response
from werkzeug.security import generate_password_hash, check_password_hash
from jinja2 import Template
import sqlite3, json, os, datetime, secrets, math
from urllib.parse import quote_plus
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# Database location.
#
# Background: Railway (and most modern PaaS — Heroku, Render, Fly.io,
# Cloud Run) runs containers with an EPHEMERAL filesystem. Anything written
# to disk inside the container is lost on every redeploy, daily container
# recycle, or scale-from-zero. Before this fix, usage.db lived next to
# app.py inside the container, so every restart wiped users, sessions,
# manual properties, hidden properties, uploaded CSV data, shortlists, the
# SDA market data, and the SDA supply snapshot.
#
# Fix: read the path from DB_PATH so it can be pointed at a mounted Railway
# Volume (which IS persistent across deploys/restarts). To migrate without
# data loss:
#   1. In Railway, attach a Volume to the service (Settings → Volumes).
#      Pick any mount path, e.g. /data.
#   2. Set the DB_PATH environment variable on the service to that path
#      plus a filename, e.g. /data/usage.db.
#   3. Redeploy. On first run init_db() creates the schema in the new
#      location. (If you had existing data worth keeping you would copy
#      usage.db onto the volume BEFORE the redeploy via a one-off shell
#      session — see RAILWAY_VOLUME_SETUP.md.)
#
# Local dev keeps working: with no DB_PATH set, it falls back to the
# repo-local usage.db exactly as before.
DB = os.environ.get('DB_PATH') or os.path.join(os.path.dirname(__file__), 'usage.db')

# Make sure the parent directory exists. Railway Volumes are mounted by
# the platform so the directory will already exist there, but creating it
# defensively keeps local-dev and other deploy targets working when the
# user points DB_PATH at a fresh location.
_db_dir = os.path.dirname(os.path.abspath(DB))
if _db_dir and not os.path.exists(_db_dir):
    try:
        os.makedirs(_db_dir, exist_ok=True)
    except OSError:
        pass  # If we can't create it, the sqlite3.connect below will surface a clearer error

# How often (in seconds) the login_required decorator refreshes a session's
# last_active timestamp. Previously this fired on every single request, which
# caused "database is locked" errors under multi-user load on SQLite. 60s of
# precision is plenty for "is this user currently active" — the admin Team
# activity view shows times rounded to the minute anyway.
SESSION_TOUCH_SECONDS = 60

# Google Maps Platform API key — read from environment (set on Railway).
# Used for: Geocoding, Places (New) Nearby Search, and Maps Static API.
GOOGLE_MAPS_API_KEY = os.environ.get('GOOGLE_MAPS_API_KEY', '')

def get_db():
    # timeout=10 makes Python wait up to 10s for the file lock before raising
    # OperationalError. busy_timeout sets the same at the SQLite engine level.
    # WAL mode (persisted to the DB file on first connect) lets readers and a
    # single writer work concurrently without blocking each other; this is the
    # main fix for the "database is locked" errors we were seeing under
    # multi-user load with the default rollback-journal mode.
    db = sqlite3.connect(DB, timeout=10.0)
    db.row_factory = sqlite3.Row
    db.execute('PRAGMA busy_timeout=5000')
    db.execute('PRAGMA journal_mode=WAL')
    db.execute('PRAGMA synchronous=NORMAL')
    return db

def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            full_name TEXT NOT NULL,
            role TEXT DEFAULT 'team',
            active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            session_token TEXT UNIQUE NOT NULL,
            login_at TEXT DEFAULT (datetime('now')),
            logout_at TEXT,
            last_active TEXT DEFAULT (datetime('now')),
            ip_address TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            event_type TEXT NOT NULL,
            event_data TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (session_id) REFERENCES sessions(id),
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS property_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            state TEXT UNIQUE NOT NULL,
            csv_data TEXT NOT NULL,
            row_count INTEGER DEFAULT 0,
            uploaded_by TEXT,
            uploaded_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS shortlists (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER UNIQUE NOT NULL,
            property_ids TEXT DEFAULT '[]',
            dd_data TEXT DEFAULT '{}',
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS config (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS sda_market (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            layer TEXT NOT NULL,           -- 'radius' | 'existing' | 'permitted'
            state TEXT NOT NULL,           -- vic | nsw | qld
            address TEXT NOT NULL,
            suburb TEXT,
            postcode TEXT,
            area_m2 REAL,
            frontage_m REAL,
            planning_zones TEXT,
            notes TEXT,
            lat REAL,
            lng REAL,
            uploaded_by TEXT,
            uploaded_at TEXT DEFAULT (datetime('now'))
        );
        CREATE INDEX IF NOT EXISTS idx_sda_market_layer ON sda_market(layer);
        CREATE INDEX IF NOT EXISTS idx_sda_market_state ON sda_market(state);
        -- Soft-delete: any logged-in user can hide a property with a required reason.
        -- Admins can review the list and restore (delete this row) or hard-delete
        -- (remove from property_data CSV blob entirely). Non-admins may un-hide
        -- only their own hide within HIDE_UNDO_SECONDS of creating it.
        CREATE TABLE IF NOT EXISTS hidden_properties (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            state TEXT NOT NULL,
            address TEXT NOT NULL,
            address_key TEXT NOT NULL,        -- lowercased + trimmed address for matching
            hidden_by_id INTEGER NOT NULL,
            hidden_by_name TEXT NOT NULL,
            hidden_at TEXT DEFAULT (datetime('now')),
            reason TEXT NOT NULL,
            FOREIGN KEY (hidden_by_id) REFERENCES users(id)
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_hidden_properties_unique
            ON hidden_properties(state, address_key);
        -- C2: manually-added properties (separate from CSV blobs in
        -- property_data). The screener pipeline treats these as first-class
        -- rows once they are merged client-side. Address is the matching
        -- key shared with hidden_properties so C1 hide/restore endpoints
        -- work on them with no extra plumbing.
        CREATE TABLE IF NOT EXISTS manual_properties (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            state TEXT NOT NULL,                 -- vic | nsw | qld
            address TEXT NOT NULL,
            address_key TEXT NOT NULL,           -- normalised for matching
            suburb TEXT NOT NULL,
            postcode TEXT,
            price REAL NOT NULL,
            land REAL NOT NULL,
            design_type TEXT NOT NULL,           -- 2p | 3p | dp
            sale_type TEXT NOT NULL,             -- Private Sale | Auction | Off-market | EOI
            auction_date TEXT,                   -- only when sale_type=Auction
            sa4 TEXT,
            sa3 TEXT,
            lat REAL,
            lng REAL,
            hosp1_name TEXT, hosp1_km REAL,
            hosp2_name TEXT, hosp2_km REAL,
            hosp3_name TEXT, hosp3_km REAL,
            medical_name TEXT, medical_km REAL,
            shop_name TEXT, shop_km REAL,
            train_name TEXT, train_km REAL,
            bus_km REAL,
            tram_name TEXT, tram_km REAL,
            created_by_id INTEGER NOT NULL,
            created_by_name TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            last_edited_by_id INTEGER,
            last_edited_by_name TEXT,
            last_edited_at TEXT,
            FOREIGN KEY (created_by_id) REFERENCES users(id),
            FOREIGN KEY (last_edited_by_id) REFERENCES users(id)
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_manual_properties_unique
            ON manual_properties(state, address_key);
        CREATE INDEX IF NOT EXISTS idx_manual_properties_state
            ON manual_properties(state);
        -- C3: SA2-level SDA supply & demand snapshot
        -- One row per (state, SA2). data_json holds the full aggregation:
        --   counts, deficit, provider lists, design-category mix.
        -- The full table is sent to the client on dashboard load so badge
        -- rendering and panel display happen instantly without per-row
        -- API calls. The blob is keyed off (state, SA2) because the same
        -- SA2 name CAN repeat across states (e.g. "Wellington" in NSW vs VIC).
        CREATE TABLE IF NOT EXISTS sda_supply (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            state TEXT NOT NULL,            -- vic | nsw | qld
            sa2 TEXT NOT NULL,
            sa3 TEXT,
            sa4 TEXT,
            data_json TEXT NOT NULL,        -- aggregated supply blob
            data_as_of TEXT,                -- date from the source files (e.g. 2025-12-31)
            uploaded_by TEXT,
            uploaded_at TEXT DEFAULT (datetime('now'))
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_sda_supply_unique
            ON sda_supply(state, sa2);
    """)
    db.commit()
    existing = db.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    if existing == 0:
        db.execute('INSERT INTO users (username,password_hash,full_name,role) VALUES (?,?,?,?)',
                   ('admin', generate_password_hash('admin123'), 'Administrator', 'admin'))
        db.commit()
    db.close()

init_db()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = session.get('token')
        if not token:
            return redirect(url_for('login'))
        # One DB connection for both the read and the (throttled) write.
        # Previously this used two separate connections per request which
        # multiplied lock contention under concurrent users.
        db = get_db()
        try:
            sess = db.execute(
                'SELECT s.*,u.username,u.full_name,u.role,u.active FROM sessions s '
                'JOIN users u ON s.user_id=u.id WHERE s.session_token=? AND s.logout_at IS NULL', (token,)
            ).fetchone()
            if not sess or not sess['active']:
                session.clear()
                return redirect(url_for('login'))
            # Throttle the last_active write to once per SESSION_TOUCH_SECONDS.
            # Refreshing the timestamp on every single request was the main
            # cause of "database is locked" errors — every page load, every
            # api/log POST, every api/me poll was queueing an UPDATE on the
            # sessions table. Sub-minute precision on last_active isn't
            # operationally useful, so we now only write when it's stale.
            # Wrapped in try/except so a transient lock here never 500s the
            # actual page the user is trying to view.
            try:
                now = datetime.datetime.utcnow()
                last_active_str = sess['last_active']
                stale = True
                if last_active_str:
                    try:
                        last_active = datetime.datetime.fromisoformat(last_active_str)
                        stale = (now - last_active).total_seconds() >= SESSION_TOUCH_SECONDS
                    except (ValueError, TypeError):
                        stale = True
                if stale:
                    db.execute('UPDATE sessions SET last_active=? WHERE session_token=?',
                               (now.isoformat(), token))
                    db.commit()
            except sqlite3.OperationalError as e:
                # DB locked / busy / etc — last_active refresh is non-critical,
                # log it and continue. Without this guard the lock would
                # propagate as a 500 on every request.
                app.logger.warning('session last_active refresh skipped: %s', e)
        finally:
            db.close()
        request.current_user = dict(sess)
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if request.current_user.get('role') != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated

def log_event(event_type, event_data=None):
    token = session.get('token')
    if not token: return
    db = get_db()
    sess = db.execute('SELECT id,user_id FROM sessions WHERE session_token=?', (token,)).fetchone()
    if sess:
        db.execute('INSERT INTO events (session_id,user_id,event_type,event_data) VALUES (?,?,?,?)',
                   (sess['id'], sess['user_id'], event_type, json.dumps(event_data) if event_data else None))
        db.commit()
    db.close()

def get_dashboard_html():
    # Serve the repo file FIRST so that committing dashboard.html to GitHub
    # actually updates the live dashboard on redeploy (this is the real deploy
    # path). Fall back to a copy stored in the DB config (legacy path, no longer
    # has an upload UI), then to a placeholder.
    p = os.path.join(os.path.dirname(__file__), 'dashboard.html')
    if os.path.exists(p):
        with open(p) as f:
            return f.read()
    db = get_db()
    row = db.execute("SELECT value FROM config WHERE key='dashboard_html'").fetchone()
    db.close()
    if row and row['value']:
        return row['value']
    return '<h1>Dashboard not loaded. Upload via admin panel.</h1>'


# ===========================================================================
# Soft-delete (hide) helpers
# ---------------------------------------------------------------------------
# Any logged-in user can hide a property from the screener by giving a reason.
# Admins can review the hidden list and restore (un-hide) or hard-delete
# (permanently remove the row from the CSV blob in property_data).
# Non-admins can un-hide only their own hide within HIDE_UNDO_SECONDS of
# creating it (powers the "Undo" toast immediately after hiding by mistake).
# ===========================================================================
HIDE_UNDO_SECONDS = 300  # 5 minutes — covers the 8-second undo toast plus slack

def _address_key(addr):
    """Normalise an address into a stable matching key.

    The CSV has free-text addresses with arbitrary whitespace, quoting, and
    case. We lower-case, collapse internal whitespace, and trim. Anything
    fancier (postcode-stripping, fuzzy match) would risk false positives;
    keep it strict so 'hide' only hits the exact row the user clicked."""
    if not addr:
        return ''
    return ' '.join(str(addr).strip().lower().split())

def _get_hidden_address_keys(state):
    """Return a set of normalised address keys hidden for the given state."""
    db = get_db()
    rows = db.execute(
        'SELECT address_key FROM hidden_properties WHERE state=?', (state,)
    ).fetchall()
    db.close()
    return {r['address_key'] for r in rows}

def _strip_hidden_from_csv(csv_data, hidden_keys, addr_col_candidates=('address', 'street address', 'street')):
    """Remove rows from a CSV blob whose Address column matches a hidden key.

    Mirrors the column-detection logic the dashboard uses client-side
    (processText in dashboard.html). Returns (new_csv_text, removed_count)."""
    if not csv_data or not hidden_keys:
        return csv_data, 0
    lines = csv_data.split('\n')
    if not lines:
        return csv_data, 0
    # Find header line (first non-empty line)
    header_idx = 0
    while header_idx < len(lines) and not lines[header_idx].strip():
        header_idx += 1
    if header_idx >= len(lines):
        return csv_data, 0
    header = lines[header_idx]
    cols = [c.strip().strip('"').lower() for c in header.split(',')]
    addr_idx = -1
    for cand in addr_col_candidates:
        for i, col in enumerate(cols):
            if col == cand or cand in col:
                addr_idx = i
                break
        if addr_idx >= 0:
            break
    if addr_idx < 0:
        # Couldn't find address column — leave CSV untouched rather than
        # silently dropping rows.
        return csv_data, 0
    out_lines = lines[:header_idx + 1]
    removed = 0
    for line in lines[header_idx + 1:]:
        if not line.strip():
            out_lines.append(line)
            continue
        # Same naive CSV-with-quoted-fields parse the dashboard uses
        cells, cur, in_q = [], '', False
        for ch in line:
            if ch == '"':
                in_q = not in_q
            elif ch == ',' and not in_q:
                cells.append(cur.strip())
                cur = ''
            else:
                cur += ch
        cells.append(cur.strip())
        addr_val = cells[addr_idx] if addr_idx < len(cells) else ''
        addr_val = addr_val.replace('"', '')
        if _address_key(addr_val) in hidden_keys:
            removed += 1
            continue
        out_lines.append(line)
    return '\n'.join(out_lines), removed


# ===========================================================================
# Google Maps proxy helpers
# ---------------------------------------------------------------------------
# These run server-side so the Google API key is never exposed to the browser
# and HTTP-referer key restrictions don't apply (since requests originate from
# Railway's server IP, not from a browser).
# ===========================================================================

def _http_get_json(url, timeout=10):
    """GET a URL and return parsed JSON. Returns dict with 'error' key on failure."""
    try:
        req = Request(url, headers={'User-Agent': 'SDA-Screener/1.0'})
        with urlopen(req, timeout=timeout) as r:
            return json.loads(r.read().decode('utf-8'))
    except HTTPError as e:
        return {'error': f'HTTP {e.code}: {e.reason}'}
    except URLError as e:
        return {'error': f'Network: {e.reason}'}
    except Exception as e:
        return {'error': str(e)}

def _http_post_json(url, body, headers=None, timeout=10):
    """POST JSON body and return parsed JSON. Returns dict with 'error' key on failure."""
    try:
        data = json.dumps(body).encode('utf-8')
        req_headers = {'Content-Type': 'application/json', 'User-Agent': 'SDA-Screener/1.0'}
        if headers:
            req_headers.update(headers)
        req = Request(url, data=data, headers=req_headers, method='POST')
        with urlopen(req, timeout=timeout) as r:
            return json.loads(r.read().decode('utf-8'))
    except HTTPError as e:
        body_text = ''
        try:
            body_text = e.read().decode('utf-8')[:500]
        except Exception:
            pass
        return {'error': f'HTTP {e.code}: {e.reason}', 'body': body_text}
    except URLError as e:
        return {'error': f'Network: {e.reason}'}
    except Exception as e:
        return {'error': str(e)}

def _haversine_km(lat1, lng1, lat2, lng2):
    """Straight-line distance in km between two lat/lng points."""
    R = 6371.0
    rlat1, rlat2 = math.radians(lat1), math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = math.sin(dlat/2)**2 + math.cos(rlat1) * math.cos(rlat2) * math.sin(dlng/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

@app.route('/api/geocode')
@login_required
def api_geocode():
    """Geocode an address to lat/lng. Returns {lat, lng} or {error}."""
    if not GOOGLE_MAPS_API_KEY:
        return jsonify({'error': 'GOOGLE_MAPS_API_KEY not configured on server'}), 500
    addr = (request.args.get('address') or '').strip()
    if not addr:
        return jsonify({'error': 'Missing address parameter'}), 400
    url = ('https://maps.googleapis.com/maps/api/geocode/json'
           '?address=' + quote_plus(addr + ', Australia') +
           '&key=' + GOOGLE_MAPS_API_KEY)
    data = _http_get_json(url)
    if 'error' in data:
        return jsonify({'error': data['error']}), 502
    if data.get('status') != 'OK' or not data.get('results'):
        return jsonify({'error': 'Geocode status: ' + data.get('status', 'unknown'),
                        'message': data.get('error_message', '')}), 502
    loc = data['results'][0]['geometry']['location']
    return jsonify({'lat': loc['lat'], 'lng': loc['lng']})

def _places_search(lat, lng, place_type, radius=5000):
    """Search Google Places near (lat, lng) for `place_type`.

    Encapsulates the hospital text-search + name-filter logic, the
    supermarket major-chain filter, and the shopping_mall name filter so the
    same rules apply to both /api/places-nearby (client-driven) and manual
    property creation (server-driven, in /api/manual-properties).

    Returns a list of dicts [{'name': str, 'km': float}, ...] sorted by
    distance ascending, or an empty list. Raises ValueError with a message
    on upstream errors so callers can decide how to surface them.
    """
    radius = max(50, min(int(radius), 50000))

    # ----- HOSPITAL: use Text Search instead of Nearby Search -----
    # Google's Nearby Search tags too many things as primaryType=hospital
    # (medical centres, kidney clinics, individual doctors). With distance
    # ranking, those false positives get ranked above real hospitals further
    # out. Text Search "hospital near {lat,lng}" understands the human
    # meaning of 'hospital' and ranks real hospitals (Mercy, Northern,
    # Royal Melbourne, etc.) at the top.
    if place_type == 'hospital':
        text_body = {
            'textQuery': 'hospital',
            'maxResultCount': 20,
            'locationBias': {
                'circle': {
                    'center': {'latitude': lat, 'longitude': lng},
                    'radius': radius
                }
            }
        }
        text_headers = {
            'X-Goog-Api-Key': GOOGLE_MAPS_API_KEY,
            'X-Goog-FieldMask': 'places.displayName,places.location,places.primaryType,places.types'
        }
        data = _http_post_json('https://places.googleapis.com/v1/places:searchText',
                               text_body, headers=text_headers)
    else:
        body = {
            'includedTypes': [place_type],
            'maxResultCount': 20,
            'rankPreference': 'DISTANCE',
            'locationRestriction': {
                'circle': {
                    'center': {'latitude': lat, 'longitude': lng},
                    'radius': radius
                }
            }
        }
        headers = {
            'X-Goog-Api-Key': GOOGLE_MAPS_API_KEY,
            'X-Goog-FieldMask': 'places.displayName,places.location,places.primaryType,places.types'
        }
        data = _http_post_json('https://places.googleapis.com/v1/places:searchNearby',
                               body, headers=headers)
    if 'error' in data:
        raise ValueError(data['error'])
    places = data.get('places') or []
    if not places:
        return []

    # For 'hospital' searches via Text Search: Google's results are already
    # high-quality (real hospitals at the top). We still apply name-based
    # filters to catch the rare bad result, then re-sort by distance.
    if place_type == 'hospital':
        # Things to exclude even when name contains 'hospital' (animal/private/etc.)
        NAME_EXCLUDE_KEYWORDS = [
            'animal hospital', 'pet hospital', 'veterinary',
            'institute', 'foundation', 'association', 'society',
            'training', 'engineering', 'consulting', 'consultancy',
            'cosmetic', 'plastic surgery'
        ]
        # Acceptable hospital naming patterns
        REQUIRED_NAME_KEYWORDS = [
            'hospital', 'health service', 'health network',
            'northern health', 'austin health', 'eastern health',
            'monash health', 'western health', 'alfred health',
            'mercy health', 'st vincent', 'royal melbourne', 'royal childrens',
            'royal children', 'box hill', 'epworth', 'cabrini', 'peter maccallum'
        ]
        EXCLUDE_TYPES = {
            'doctor', 'medical_lab', 'dental_clinic', 'dentist',
            'physiotherapist', 'chiropractor', 'pharmacy', 'drugstore',
            'veterinary_care', 'spa', 'beauty_salon', 'wellness_center'
        }
        def name_required(nm):
            n = (nm or '').lower()
            return any(kw in n for kw in REQUIRED_NAME_KEYWORDS)
        def name_excluded(nm):
            n = (nm or '').lower()
            return any(kw in n for kw in NAME_EXCLUDE_KEYWORDS)

        filtered = []
        for p in places:
            types = set(p.get('types') or [])
            nm = ((p.get('displayName') or {}).get('text')) or ''
            if (name_required(nm)
                    and not name_excluded(nm)
                    and not (types & EXCLUDE_TYPES)):
                filtered.append(p)
        places = filtered

    # For 'supermarket': prefer well-known major retailers by name.
    if place_type == 'supermarket':
        SUPERMARKET_KEYWORDS = [
            'coles', 'woolworths', 'woolies', 'aldi', 'iga', 'foodworks',
            'costco', 'drakes', 'spudshed', 'harris farm', 'ritchies'
        ]
        def is_known_chain(nm):
            n = (nm or '').lower()
            return any(kw in n for kw in SUPERMARKET_KEYWORDS)
        places = [p for p in places
                  if is_known_chain(((p.get('displayName') or {}).get('text')) or '')]

    # For 'shopping_mall': require the name to contain a shopping-centre keyword.
    if place_type == 'shopping_mall':
        MALL_KEYWORDS = [
            'shopping centre', 'shopping center', 'shopping mall',
            'shopping village', 'shopping plaza', 'plaza', 'westfield',
            'marketplace', 'arcade', 'town centre', 'town center',
            'square', 'mall', 'qv ', 'emporium', 'galleria', 'outlet'
        ]
        def is_mall(nm):
            n = (nm or '').lower()
            return any(kw in n for kw in MALL_KEYWORDS)
        places = [p for p in places
                  if is_mall(((p.get('displayName') or {}).get('text')) or '')]

    out = []
    for p in places:
        loc = p.get('location') or {}
        if 'latitude' not in loc or 'longitude' not in loc:
            continue
        nm = ((p.get('displayName') or {}).get('text')) or '—'
        km = _haversine_km(lat, lng, loc['latitude'], loc['longitude'])
        out.append({'name': nm, 'km': round(km, 2)})
    out.sort(key=lambda x: x['km'])
    return out


def _compute_amenities(lat, lng):
    """Run the six Google Places lookups used for a manual property and
    return a flat dict matching the column names the dashboard expects:
    hosp1_name/km, hosp2_name/km, hosp3_name/km, medical_name/km,
    shop_name/km, train_name/km, bus_km, tram_name/km. Any field that
    can't be filled comes back as ''. Errors from the upstream API are
    logged but never raise — a property still saves even if a lookup
    fails."""
    out = {
        'hosp1_name': '', 'hosp1_km': '',
        'hosp2_name': '', 'hosp2_km': '',
        'hosp3_name': '', 'hosp3_km': '',
        'medical_name': '', 'medical_km': '',
        'shop_name': '', 'shop_km': '',
        'train_name': '', 'train_km': '',
        'bus_km': '',
        'tram_name': '', 'tram_km': '',
    }
    if not GOOGLE_MAPS_API_KEY:
        return out

    def safe(place_type, radius=5000):
        try:
            return _places_search(lat, lng, place_type, radius)
        except ValueError as e:
            print('[_compute_amenities]', place_type, 'failed:', e, flush=True)
            return []

    # Hospitals: top 3
    hosps = safe('hospital', radius=15000)  # wider radius — hospitals are rarer
    for i, h in enumerate(hosps[:3], start=1):
        out['hosp' + str(i) + '_name'] = h['name']
        out['hosp' + str(i) + '_km'] = h['km']

    # Doctor (any GP / medical centre)
    docs = safe('doctor')
    if docs:
        out['medical_name'] = docs[0]['name']
        out['medical_km'] = docs[0]['km']

    # Supermarket — fall back to shopping_mall when no major chain nearby
    sups = safe('supermarket')
    if sups:
        out['shop_name'] = sups[0]['name']
        out['shop_km'] = sups[0]['km']
    else:
        malls = safe('shopping_mall')
        if malls:
            out['shop_name'] = malls[0]['name']
            out['shop_km'] = malls[0]['km']

    # Train
    trains = safe('train_station')
    if trains:
        out['train_name'] = trains[0]['name']
        out['train_km'] = trains[0]['km']

    # Bus — store distance only (matches CSV schema, which has bus_km but no bus_name)
    buses = safe('bus_station', radius=2000)
    if buses:
        out['bus_km'] = buses[0]['km']

    # Tram
    trams = safe('light_rail_station', radius=3000)
    if trams:
        out['tram_name'] = trams[0]['name']
        out['tram_km'] = trams[0]['km']

    return out


@app.route('/api/places-nearby')
@login_required
def api_places_nearby():
    """Find the nearest place of a given type. Args: lat, lng, type, radius (m).
    Returns {name, km} of the closest match, or {error} / {name: null} if none found."""
    if not GOOGLE_MAPS_API_KEY:
        return jsonify({'error': 'GOOGLE_MAPS_API_KEY not configured on server'}), 500
    try:
        lat = float(request.args.get('lat', ''))
        lng = float(request.args.get('lng', ''))
    except ValueError:
        return jsonify({'error': 'Invalid lat/lng'}), 400
    place_type = (request.args.get('type') or '').strip()
    if not place_type:
        return jsonify({'error': 'Missing type'}), 400
    try:
        radius = int(request.args.get('radius', 5000))
    except ValueError:
        radius = 5000

    try:
        results = _places_search(lat, lng, place_type, radius=radius)
    except ValueError as e:
        return jsonify({'error': str(e)}), 502

    if not results:
        return jsonify({'name': None, 'km': None})
    return jsonify({'name': results[0]['name'], 'km': results[0]['km']})



@app.route('/api/static-map')
@login_required
def api_static_map():
    """Proxy a Google Static Maps request. Args: lat, lng (or address as fallback).
    Returns the PNG image bytes directly."""
    if not GOOGLE_MAPS_API_KEY:
        return jsonify({'error': 'GOOGLE_MAPS_API_KEY not configured on server'}), 500
    lat = request.args.get('lat')
    lng = request.args.get('lng')
    addr = request.args.get('address')
    if lat and lng:
        try:
            float(lat); float(lng)
        except ValueError:
            return jsonify({'error': 'Invalid lat/lng'}), 400
        center = lat + ',' + lng
    elif addr:
        center = quote_plus(addr.strip() + ', Australia')
    else:
        return jsonify({'error': 'Missing lat/lng or address'}), 400
    try:
        zoom = int(request.args.get('zoom', 14))
    except ValueError:
        zoom = 14
    zoom = max(1, min(zoom, 20))
    size = request.args.get('size', '600x300')
    if not all(p.isdigit() for p in size.split('x')) or 'x' not in size:
        size = '600x300'
    url = ('https://maps.googleapis.com/maps/api/staticmap'
           '?center=' + center +
           '&zoom=' + str(zoom) +
           '&size=' + size +
           '&scale=2&maptype=roadmap'
           '&markers=color:red%7Csize:mid%7C' + center +
           '&key=' + GOOGLE_MAPS_API_KEY)
    try:
        req = Request(url, headers={'User-Agent': 'SDA-Screener/1.0'})
        with urlopen(req, timeout=15) as r:
            img_bytes = r.read()
            content_type = r.headers.get('Content-Type', 'image/png')
        resp = make_response(img_bytes)
        resp.headers['Content-Type'] = content_type
        resp.headers['Cache-Control'] = 'private, max-age=3600'
        return resp
    except (HTTPError, URLError) as e:
        return jsonify({'error': str(e)}), 502

@app.route('/api/suburb-info')
@login_required
def api_suburb_info():
    """Look up suburb info from Wikipedia.
    Query params: suburb, state (e.g. 'Thornbury', 'VIC')
    Returns: {summary: str, population: int|null, title: str} or {error}
    No API key needed for Wikipedia REST API."""
    suburb = (request.args.get('suburb') or '').strip()
    state_code = (request.args.get('state') or '').strip().upper()
    if not suburb:
        return jsonify({'error': 'Missing suburb parameter'}), 400
    state_full = {'VIC': 'Victoria', 'NSW': 'New South Wales', 'QLD': 'Queensland',
                  'SA': 'South Australia', 'WA': 'Western Australia',
                  'TAS': 'Tasmania', 'NT': 'Northern Territory', 'ACT': 'Australian Capital Territory'}
    # Try increasingly specific queries, returning first match
    titles_to_try = []
    if state_code in state_full:
        titles_to_try.append(suburb + ', ' + state_full[state_code])
    titles_to_try.append(suburb)
    summary, title_used = None, None
    for title in titles_to_try:
        url = ('https://en.wikipedia.org/api/rest_v1/page/summary/'
               + quote_plus(title.replace(' ', '_')))
        try:
            req = Request(url, headers={'User-Agent': 'SDA-Screener/1.0 (radius project mgmt)'})
            with urlopen(req, timeout=8) as r:
                data = json.loads(r.read().decode('utf-8'))
            # Skip disambiguation pages and missing pages
            if data.get('type') == 'disambiguation' or data.get('type') == 'no-extract':
                continue
            extract = (data.get('extract') or '').strip()
            if extract:
                summary = extract
                title_used = data.get('title') or title
                break
        except (HTTPError, URLError, ValueError):
            continue
    if not summary:
        return jsonify({'error': 'Suburb info not found', 'suburb': suburb})
    # Try to extract a population figure from the summary text
    population = None
    import re as _re
    # Match "population of 21,567" or "21,567 people" or "(2021 population: 21,567)"
    patterns = [
        r'population[^\d]{0,30}(\d{1,3}(?:,\d{3})+|\d{4,7})',
        r'(\d{1,3}(?:,\d{3})+|\d{4,7})\s*(?:people|residents|inhabitants)',
    ]
    for pat in patterns:
        m = _re.search(pat, summary, _re.IGNORECASE)
        if m:
            try:
                population = int(m.group(1).replace(',', ''))
                break
            except ValueError:
                pass
    # Trim summary to ~3 sentences for compact display
    sentences = _re.split(r'(?<=[.!?])\s+', summary)
    short_summary = ' '.join(sentences[:3]).strip()
    if len(short_summary) > 450:
        short_summary = short_summary[:447].rstrip() + '...'
    return jsonify({
        'title': title_used,
        'summary': short_summary,
        'population': population
    })

LOGIN_HTML = """<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>SDA Property Screener - Login</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Arial,sans-serif;background:linear-gradient(135deg,#001540 0%,#002060 50%,#003090 100%);min-height:100vh;display:flex;align-items:center;justify-content:center}
.card{background:#fff;border-radius:16px;padding:48px 40px;width:100%;max-width:400px;box-shadow:0 24px 80px rgba(0,0,0,0.4)}
.logo{text-align:center;margin-bottom:32px}
.logo-icon{font-size:40px;margin-bottom:12px}
.logo h1{font-size:22px;font-weight:800;color:#002060;margin-bottom:4px}
.logo p{font-size:13px;color:#6B7280}
label{display:block;font-size:12px;font-weight:700;color:#374151;margin-bottom:6px;text-transform:uppercase}
input{width:100%;padding:12px 14px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;margin-bottom:16px}
input:focus{outline:none;border-color:#185FA5}
button{width:100%;padding:14px;background:#002060;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:700;cursor:pointer}
button:hover{background:#003090}
.error{background:#fff3f3;border:1px solid #fca5a5;color:#A32D2D;padding:12px 14px;border-radius:8px;font-size:13px;margin-bottom:16px}
.footer{text-align:center;margin-top:24px;font-size:12px;color:#9CA3AF}
</style></head>
<body><div class="card">
<div class="logo"><div class="logo-icon">&#127968;</div><h1>SDA Property Screener</h1><p>NDIS Investment Analysis Platform</p></div>
{% if error %}<div class="error">{{ error }}</div>{% endif %}
<form method="POST"><label>USERNAME</label><input name="username" placeholder="Enter your username" autofocus>
<label>PASSWORD</label><input type="password" name="password" placeholder="Enter your password">
<button type="submit">Sign In</button></form>
<div class="footer">Confidential - Authorised users only</div>
</div></body></html>"""

@app.route('/')
def index():
    if 'token' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username','').strip().lower()
        password = request.form.get('password','')
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE username=? AND active=1', (username,)).fetchone()
        db.close()
        if user and check_password_hash(user['password_hash'], password):
            token = secrets.token_urlsafe(32)
            db = get_db()
            db.execute('INSERT INTO sessions (user_id,session_token,ip_address) VALUES (?,?,?)',
                       (user['id'], token, request.remote_addr))
            db.commit()
            db.close()
            session['token'] = token
            log_event('login', {'username': username})
            return redirect(url_for('dashboard'))
        error = 'Invalid username or password.'
    return make_response(Template(LOGIN_HTML).render(error=error))

@app.route('/logout')
def logout():
    token = session.get('token')
    if token:
        log_event('logout')
        db = get_db()
        db.execute('UPDATE sessions SET logout_at=? WHERE session_token=?',
                   (datetime.datetime.utcnow().isoformat(), token))
        db.commit()
        db.close()
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    log_event('dashboard_view')
    html = get_dashboard_html()
    # Inject the Maps JS API key for the SDA Map tab (browser-side rendering).
    # The same GOOGLE_MAPS_API_KEY env var is reused — for browser use, ensure
    # Application restrictions in Google Cloud Console include HTTP referrer
    # for your Railway domain, AND that Maps JavaScript API is enabled.
    html = html.replace('{{GOOGLE_MAPS_API_KEY}}', GOOGLE_MAPS_API_KEY or '')
    return make_response(html)

@app.route('/api/log', methods=['POST'])
@login_required
def api_log():
    data = request.get_json()
    if data and data.get('type'):
        log_event(data['type'], data.get('data'))
    return jsonify({'ok': True})

@app.route('/api/properties/<state>')
@login_required
def get_properties(state):
    if state not in ('vic','nsw','qld'): abort(400)
    db = get_db()
    row = db.execute('SELECT csv_data,row_count,uploaded_at FROM property_data WHERE state=?', (state,)).fetchone()
    db.close()
    if not row: return jsonify({'data': None, 'rows': 0, 'uploaded_at': None})
    csv_data = row['csv_data']
    row_count = row['row_count']
    # Strip soft-deleted (hidden) rows before sending to the dashboard.
    # Admins reviewing the hidden list use /api/hidden-properties instead.
    hidden_keys = _get_hidden_address_keys(state)
    if hidden_keys:
        csv_data, removed = _strip_hidden_from_csv(csv_data, hidden_keys)
        row_count = max(0, row_count - removed)
    return jsonify({'data': csv_data, 'rows': row_count, 'uploaded_at': row['uploaded_at']})

@app.route('/api/properties/status')
@login_required
def properties_status():
    db = get_db()
    rows = db.execute('SELECT state,row_count,uploaded_at,uploaded_by FROM property_data').fetchall()
    db.close()
    return jsonify({r['state']: {'rows': r['row_count'], 'uploaded_at': r['uploaded_at'], 'uploaded_by': r['uploaded_by']} for r in rows})

@app.route('/api/shortlist', methods=['GET'])
@login_required
def get_shortlist():
    uid = request.current_user['id']
    db = get_db()
    row = db.execute('SELECT property_ids,dd_data FROM shortlists WHERE user_id=?', (uid,)).fetchone()
    db.close()
    if not row: return jsonify({'ids': [], 'dd': {}})
    return jsonify({'ids': json.loads(row['property_ids']), 'dd': json.loads(row['dd_data'] or '{}')})

@app.route('/api/shortlist', methods=['POST'])
@login_required
def save_shortlist():
    uid = request.current_user['id']
    data = request.get_json()
    ids = json.dumps(data.get('ids', []))
    dd = json.dumps(data.get('dd', {}))
    db = get_db()
    db.execute("""INSERT INTO shortlists (user_id,property_ids,dd_data) VALUES (?,?,?)
                  ON CONFLICT(user_id) DO UPDATE SET property_ids=excluded.property_ids,
                  dd_data=excluded.dd_data,updated_at=datetime('now')""", (uid, ids, dd))
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ===========================================================================
# Soft-delete (hide) API endpoints
# ===========================================================================

@app.route('/api/me')
@login_required
def api_me():
    """Return identity for the logged-in user.

    The dashboard uses this to decide whether to show admin-only controls
    (Restore / Hard-delete) and to label its own hide-undo eligibility."""
    u = request.current_user
    return jsonify({
        'username': u['username'],
        'full_name': u['full_name'],
        'role': u['role'],
        'is_admin': u['role'] == 'admin',
        'hide_undo_seconds': HIDE_UNDO_SECONDS,
    })


@app.route('/api/hide-property', methods=['POST'])
@login_required
def api_hide_property():
    """Soft-delete a property from the screener.

    Any logged-in user can call this. Reason is required (min 3 chars).
    Returns the new hidden_properties row id so the client can offer Undo."""
    data = request.get_json(silent=True) or {}
    state = (data.get('state') or '').strip().lower()
    address = (data.get('address') or '').strip()
    reason = (data.get('reason') or '').strip()
    if state not in ('vic', 'nsw', 'qld'):
        return jsonify({'error': 'Invalid state'}), 400
    if not address:
        return jsonify({'error': 'Missing address'}), 400
    if len(reason) < 3:
        return jsonify({'error': 'Reason is required (min 3 characters)'}), 400
    addr_key = _address_key(address)
    if not addr_key:
        return jsonify({'error': 'Address could not be normalised'}), 400
    u = request.current_user
    db = get_db()
    try:
        cur = db.execute(
            """INSERT INTO hidden_properties
               (state, address, address_key, hidden_by_id, hidden_by_name, reason)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (state, address, addr_key, u['user_id'], u['full_name'], reason)
        )
        db.commit()
        new_id = cur.lastrowid
    except sqlite3.IntegrityError:
        # Already hidden by someone else — that's fine, the property is gone
        # either way. Return the existing row id so Undo still works for the
        # admin (and the original hider) but not for this user.
        existing = db.execute(
            'SELECT id FROM hidden_properties WHERE state=? AND address_key=?',
            (state, addr_key)
        ).fetchone()
        db.close()
        return jsonify({
            'ok': True,
            'id': existing['id'] if existing else None,
            'already_hidden': True,
        })
    db.close()
    log_event('property_hidden', {'state': state, 'address': address, 'reason': reason})
    return jsonify({'ok': True, 'id': new_id, 'already_hidden': False})


@app.route('/api/unhide-property', methods=['POST'])
@login_required
def api_unhide_property():
    """Restore a hidden property (delete the hidden_properties row).

    Admins can unhide any row. Non-admins can unhide only their own hide
    within HIDE_UNDO_SECONDS — this powers the 'Undo' toast immediately
    after a mistaken click."""
    data = request.get_json(silent=True) or {}
    try:
        hid = int(data.get('id'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Missing or invalid id'}), 400
    u = request.current_user
    db = get_db()
    row = db.execute(
        'SELECT id, state, address, hidden_by_id, hidden_at FROM hidden_properties WHERE id=?',
        (hid,)
    ).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Hide record not found'}), 404
    is_admin = u['role'] == 'admin'
    is_owner = (row['hidden_by_id'] == u['user_id'])
    within_window = False
    if is_owner:
        try:
            # SQLite stores as 'YYYY-MM-DD HH:MM:SS' UTC
            hidden_at = datetime.datetime.strptime(row['hidden_at'], '%Y-%m-%d %H:%M:%S')
            age = (datetime.datetime.utcnow() - hidden_at).total_seconds()
            within_window = age <= HIDE_UNDO_SECONDS
        except (ValueError, TypeError):
            within_window = False
    if not (is_admin or (is_owner and within_window)):
        db.close()
        if is_owner:
            return jsonify({'error': 'Undo window has expired. Ask an admin to restore.'}), 403
        return jsonify({'error': 'Only an admin can restore this property'}), 403
    db.execute('DELETE FROM hidden_properties WHERE id=?', (hid,))
    db.commit()
    db.close()
    log_event('property_unhidden', {
        'state': row['state'], 'address': row['address'],
        'by_admin': is_admin and not is_owner,
    })
    return jsonify({'ok': True})


@app.route('/api/hidden-properties')
@admin_required
def api_list_hidden():
    """List all currently hidden properties. Admin only."""
    db = get_db()
    rows = db.execute(
        """SELECT id, state, address, hidden_by_id, hidden_by_name,
                  hidden_at, reason
           FROM hidden_properties ORDER BY hidden_at DESC"""
    ).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/hard-delete-property', methods=['POST'])
@admin_required
def api_hard_delete_property():
    """Permanently remove a property from the CSV blob for its state.

    Also removes any matching hidden_properties row so the address can later
    be re-added (e.g. via manual entry in Stage C2) without conflict.
    Admin only."""
    data = request.get_json(silent=True) or {}
    state = (data.get('state') or '').strip().lower()
    address = (data.get('address') or '').strip()
    if state not in ('vic', 'nsw', 'qld'):
        return jsonify({'error': 'Invalid state'}), 400
    if not address:
        return jsonify({'error': 'Missing address'}), 400
    addr_key = _address_key(address)
    db = get_db()
    row = db.execute(
        'SELECT csv_data, row_count FROM property_data WHERE state=?', (state,)
    ).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'No property data for that state'}), 404
    new_csv, removed = _strip_hidden_from_csv(row['csv_data'], {addr_key})
    if removed == 0:
        # The address wasn't in the CSV. Still purge any hidden_properties row
        # for it (so the admin can clean up orphan hides).
        db.execute(
            'DELETE FROM hidden_properties WHERE state=? AND address_key=?',
            (state, addr_key)
        )
        db.commit()
        db.close()
        return jsonify({
            'ok': True,
            'removed_from_csv': 0,
            'note': 'Address was not present in the current CSV; any hide record cleared.',
        })
    new_row_count = max(0, (row['row_count'] or 0) - removed)
    db.execute(
        'UPDATE property_data SET csv_data=?, row_count=? WHERE state=?',
        (new_csv, new_row_count, state)
    )
    db.execute(
        'DELETE FROM hidden_properties WHERE state=? AND address_key=?',
        (state, addr_key)
    )
    db.commit()
    db.close()
    log_event('property_hard_deleted', {'state': state, 'address': address, 'rows_removed': removed})
    return jsonify({'ok': True, 'removed_from_csv': removed})


# ===========================================================================
# C2: Manual property addition
# ---------------------------------------------------------------------------
# Anyone signed-in can add a property; the row is geocoded server-side and
# the same Google Places amenity lookups run that the CSV pipeline does, so
# manual entries are indistinguishable from CSV ones in the screener.
# The hide/restore endpoints from C1 already key on (state, address_key)
# and therefore work on manual properties with no further code. Hard delete
# of a manual property (admin only) ALSO purges any matching hidden_properties
# row so the address can be re-added cleanly later.
# ===========================================================================

VALID_DESIGN_TYPES = {'2p', '3p', 'dp'}
VALID_SALE_TYPES = {'Private Sale', 'Auction', 'Off-market', 'Expression of Interest'}


def _manual_row_to_dict(r):
    """Serialise a manual_properties SQLite row for the JSON API."""
    return {
        'id': r['id'],
        'state': r['state'],
        'address': r['address'],
        'suburb': r['suburb'],
        'postcode': r['postcode'] or '',
        'price': r['price'],
        'land': r['land'],
        'design_type': r['design_type'],
        'sale_type': r['sale_type'],
        'auction_date': r['auction_date'] or '',
        'sa4': r['sa4'] or '',
        'sa3': r['sa3'] or '',
        'lat': r['lat'],
        'lng': r['lng'],
        'hosp1_name': r['hosp1_name'] or '', 'hosp1_km': r['hosp1_km'] if r['hosp1_km'] is not None else '',
        'hosp2_name': r['hosp2_name'] or '', 'hosp2_km': r['hosp2_km'] if r['hosp2_km'] is not None else '',
        'hosp3_name': r['hosp3_name'] or '', 'hosp3_km': r['hosp3_km'] if r['hosp3_km'] is not None else '',
        'medical_name': r['medical_name'] or '', 'medical_km': r['medical_km'] if r['medical_km'] is not None else '',
        'shop_name': r['shop_name'] or '', 'shop_km': r['shop_km'] if r['shop_km'] is not None else '',
        'train_name': r['train_name'] or '', 'train_km': r['train_km'] if r['train_km'] is not None else '',
        'bus_km': r['bus_km'] if r['bus_km'] is not None else '',
        'tram_name': r['tram_name'] or '', 'tram_km': r['tram_km'] if r['tram_km'] is not None else '',
        'created_by_id': r['created_by_id'],
        'created_by_name': r['created_by_name'],
        'created_at': r['created_at'],
        'last_edited_by_id': r['last_edited_by_id'],
        'last_edited_by_name': r['last_edited_by_name'] or '',
        'last_edited_at': r['last_edited_at'] or '',
    }


def _validate_manual_payload(data, partial=False):
    """Validate the JSON payload for create / update.

    Returns (cleaned_dict, error_or_none). `partial=True` skips required-field
    checks (used on PUT where the address is fixed and only changed fields
    matter; required fields still get range-checked if present)."""
    out = {}

    state = (data.get('state') or '').strip().lower()
    if state or not partial:
        if state not in ('vic', 'nsw', 'qld'):
            return None, 'state must be vic, nsw or qld'
        out['state'] = state

    address = (data.get('address') or '').strip()
    if address or not partial:
        if len(address) < 5:
            return None, 'address is required (min 5 characters)'
        out['address'] = address

    suburb = (data.get('suburb') or '').strip()
    if suburb or not partial:
        if not suburb:
            return None, 'suburb is required'
        out['suburb'] = suburb

    # Postcode is optional; if given, must be 4 digits
    if 'postcode' in data:
        pc = (data.get('postcode') or '').strip()
        if pc and not (pc.isdigit() and len(pc) == 4):
            return None, 'postcode must be 4 digits'
        out['postcode'] = pc

    if 'price' in data or not partial:
        try:
            price = float(data.get('price'))
        except (TypeError, ValueError):
            return None, 'price must be a number'
        if price < 100000:
            return None, 'price must be at least $100,000'
        out['price'] = price

    if 'land' in data or not partial:
        try:
            land = float(data.get('land'))
        except (TypeError, ValueError):
            return None, 'land size must be a number'
        # Match the screener's existing applyF filter (>1000m² is excluded)
        if land <= 0 or land > 1000:
            return None, 'land size must be between 1 and 1000 m²'
        out['land'] = land

    if 'design_type' in data or not partial:
        dt = (data.get('design_type') or '').strip().lower()
        if dt not in VALID_DESIGN_TYPES:
            return None, 'design_type must be one of: 2p, 3p, dp'
        out['design_type'] = dt

    if 'sale_type' in data or not partial:
        st = (data.get('sale_type') or '').strip()
        if st not in VALID_SALE_TYPES:
            return None, 'sale_type must be one of: ' + ', '.join(sorted(VALID_SALE_TYPES))
        out['sale_type'] = st

    # auction_date — required iff effective sale_type is Auction.
    # We only run this check when the request actually touches sale_type or
    # auction_date. A partial PUT that just changes (say) price must not
    # demand the user re-supply auction_date on every edit.
    touches_auction = ('sale_type' in out) or ('auction_date' in data)
    if touches_auction:
        effective_sale = out.get('sale_type') or data.get('_existing_sale_type')
        if effective_sale == 'Auction':
            ad = (data.get('auction_date') or '').strip()
            if not ad:
                # If we're not changing sale_type (i.e. it's already Auction
                # in the DB) and the client sent an empty auction_date,
                # treat that as 'keep existing' rather than 'clear it'.
                if 'sale_type' not in out and 'auction_date' not in data:
                    pass
                else:
                    return None, 'auction_date is required when sale_type=Auction'
            else:
                try:
                    datetime.datetime.strptime(ad, '%Y-%m-%d')
                except ValueError:
                    return None, 'auction_date must be in YYYY-MM-DD format'
                out['auction_date'] = ad
        else:
            # Switching away from Auction (or non-Auction create): clear the date.
            out['auction_date'] = None

    # SA4/SA3 — accept whatever the client looked up (or the user override)
    if 'sa4' in data:
        out['sa4'] = (data.get('sa4') or '').strip()
    if 'sa3' in data:
        out['sa3'] = (data.get('sa3') or '').strip()

    return out, None


@app.route('/api/manual-properties', methods=['GET'])
@login_required
def api_list_manual_properties():
    """List manual properties. Hidden ones (per C1 hidden_properties)
    are stripped before returning, so the dashboard sees the same view it
    does for CSV rows — without that, hiding a manual property via the
    existing ✕ button would not actually remove it from the screener."""
    db = get_db()
    rows = db.execute(
        'SELECT * FROM manual_properties ORDER BY created_at DESC'
    ).fetchall()
    # Build the set of (state, address_key) pairs that are hidden so we
    # can drop them in a single pass.
    hidden = db.execute(
        'SELECT state, address_key FROM hidden_properties'
    ).fetchall()
    db.close()
    hidden_set = {(h['state'], h['address_key']) for h in hidden}
    out = []
    for r in rows:
        if (r['state'], r['address_key']) in hidden_set:
            continue
        out.append(_manual_row_to_dict(r))
    return jsonify(out)


@app.route('/api/manual-properties/_diag', methods=['GET'])
@login_required
def api_manual_properties_diag():
    """Diagnostic endpoint for investigating 'manual properties disappear
    after logout/login' reports. Returns a summary that lets us tell apart
    the three possible failure modes:

      1. Properties are missing from the DB entirely  → total_rows is low
      2. Properties are in the DB but filtered out as hidden  → visible_rows < total_rows
      3. Properties are in the DB and not hidden, but the client isn't
         displaying them  → both counts look fine here, so the bug is
         client-side and the diagnostic ring buffer in dashboard.html will
         show what fetch() actually returned.

    Returns total_rows, visible_rows, hidden_rows, last_5_created (with
    minimal fields — id, state, address, created_by, created_at), and the
    list of currently-hidden (state, address_key) pairs for cross-reference.
    """
    db = get_db()
    total = db.execute('SELECT COUNT(*) FROM manual_properties').fetchone()[0]
    hidden = db.execute(
        'SELECT state, address_key FROM hidden_properties'
    ).fetchall()
    hidden_set = {(h['state'], h['address_key']) for h in hidden}
    all_rows = db.execute(
        'SELECT state, address_key, id, address, created_by_name, created_at '
        'FROM manual_properties ORDER BY created_at DESC'
    ).fetchall()
    db.close()
    visible = sum(1 for r in all_rows if (r['state'], r['address_key']) not in hidden_set)
    return jsonify({
        'total_rows': total,
        'visible_rows': visible,
        'hidden_rows': total - visible,
        'last_5_created': [
            {
                'id': r['id'],
                'state': r['state'],
                'address': r['address'],
                'created_by': r['created_by_name'],
                'created_at': r['created_at'],
                'is_hidden': (r['state'], r['address_key']) in hidden_set,
            }
            for r in all_rows[:5]
        ],
        'hidden_pairs_total': len(hidden_set),
    })


@app.route('/api/manual-properties', methods=['POST'])
@login_required
def api_create_manual_property():
    """Create a new manual property. Geocodes the address and runs the
    Google Places amenity lookups server-side."""
    data = request.get_json(silent=True) or {}
    cleaned, err = _validate_manual_payload(data, partial=False)
    if err:
        return jsonify({'error': err}), 400

    state = cleaned['state']
    address = cleaned['address']
    addr_key = _address_key(address)

    # Reject if this address already exists as a manual property in the
    # same state (the unique index would catch it but we'd rather a clean
    # 409 than an IntegrityError).
    db = get_db()
    existing = db.execute(
        'SELECT id FROM manual_properties WHERE state=? AND address_key=?',
        (state, addr_key)
    ).fetchone()
    if existing:
        db.close()
        return jsonify({'error': 'A manual property with that address already exists in ' + state.upper(),
                        'existing_id': existing['id']}), 409

    # Geocode. We don't hard-fail on geocode failure — the amenity lookups
    # will silently come back empty and the user can edit + recompute later.
    lat, lng = None, None
    if GOOGLE_MAPS_API_KEY:
        url = ('https://maps.googleapis.com/maps/api/geocode/json'
               '?address=' + quote_plus(address + ', Australia') +
               '&key=' + GOOGLE_MAPS_API_KEY)
        geo = _http_get_json(url)
        if geo.get('status') == 'OK' and geo.get('results'):
            loc = geo['results'][0]['geometry']['location']
            lat, lng = loc['lat'], loc['lng']

    amenities = _compute_amenities(lat, lng) if (lat is not None and lng is not None) else {
        'hosp1_name': '', 'hosp1_km': '',
        'hosp2_name': '', 'hosp2_km': '',
        'hosp3_name': '', 'hosp3_km': '',
        'medical_name': '', 'medical_km': '',
        'shop_name': '', 'shop_km': '',
        'train_name': '', 'train_km': '',
        'bus_km': '',
        'tram_name': '', 'tram_km': '',
    }

    u = request.current_user
    try:
        cur = db.execute("""INSERT INTO manual_properties
            (state, address, address_key, suburb, postcode, price, land,
             design_type, sale_type, auction_date, sa4, sa3, lat, lng,
             hosp1_name, hosp1_km, hosp2_name, hosp2_km, hosp3_name, hosp3_km,
             medical_name, medical_km, shop_name, shop_km, train_name, train_km,
             bus_km, tram_name, tram_km, created_by_id, created_by_name)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (state, address, addr_key, cleaned['suburb'], cleaned.get('postcode') or None,
             cleaned['price'], cleaned['land'], cleaned['design_type'],
             cleaned['sale_type'], cleaned.get('auction_date'),
             cleaned.get('sa4') or None, cleaned.get('sa3') or None,
             lat, lng,
             amenities['hosp1_name'] or None, amenities['hosp1_km'] or None,
             amenities['hosp2_name'] or None, amenities['hosp2_km'] or None,
             amenities['hosp3_name'] or None, amenities['hosp3_km'] or None,
             amenities['medical_name'] or None, amenities['medical_km'] or None,
             amenities['shop_name'] or None, amenities['shop_km'] or None,
             amenities['train_name'] or None, amenities['train_km'] or None,
             amenities['bus_km'] or None,
             amenities['tram_name'] or None, amenities['tram_km'] or None,
             u['user_id'], u['full_name']))
        db.commit()
        new_id = cur.lastrowid
    except sqlite3.IntegrityError:
        # Race condition: another request slipped past the explicit
        # duplicate check above and inserted the same (state, address_key)
        # between our SELECT and our INSERT. Return the same clean 409 the
        # explicit path returns rather than leaking the SQLite error.
        winner = db.execute(
            'SELECT id FROM manual_properties WHERE state=? AND address_key=?',
            (state, addr_key)
        ).fetchone()
        db.close()
        return jsonify({
            'error': 'A manual property with that address already exists in ' + state.upper(),
            'existing_id': winner['id'] if winner else None,
        }), 409

    row = db.execute('SELECT * FROM manual_properties WHERE id=?', (new_id,)).fetchone()
    db.close()
    log_event('manual_property_created', {'id': new_id, 'state': state, 'address': address})
    return jsonify({'ok': True, 'property': _manual_row_to_dict(row)})


@app.route('/api/manual-properties/<int:mid>', methods=['PUT'])
@login_required
def api_update_manual_property(mid):
    """Update a manual property. Only the creator and admins can edit.
    Address cannot be changed (it's the matching key for hides). Supports
    optional ?recompute_amenities=true to redo the Google Places lookups."""
    u = request.current_user
    db = get_db()
    row = db.execute('SELECT * FROM manual_properties WHERE id=?', (mid,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Manual property not found'}), 404

    is_admin = u['role'] == 'admin'
    is_owner = (row['created_by_id'] == u['user_id'])
    if not (is_admin or is_owner):
        db.close()
        return jsonify({'error': 'Only the creator or an admin can edit this property'}), 403

    data = request.get_json(silent=True) or {}
    # Hand the validator the existing sale_type so auction_date validation
    # works even when the client only sends auction_date.
    data.setdefault('_existing_sale_type', row['sale_type'])
    cleaned, err = _validate_manual_payload(data, partial=True)
    if err:
        db.close()
        return jsonify({'error': err}), 400

    # Address is immutable — silently drop it if the client sent one.
    cleaned.pop('address', None)
    # State change is allowed (sometimes a row gets entered under the wrong
    # state); recompute address_key isn't needed because address didn't change.

    if not cleaned and 'recompute_amenities' not in request.args:
        db.close()
        return jsonify({'error': 'No editable fields supplied'}), 400

    # Build dynamic UPDATE
    set_clauses, values = [], []
    for k, v in cleaned.items():
        set_clauses.append(k + '=?')
        values.append(v)

    # Optional amenity recompute
    if request.args.get('recompute_amenities', '').lower() == 'true':
        lat = row['lat']
        lng = row['lng']
        # If lat/lng missing, try to geocode now using the (immutable) address
        if (lat is None or lng is None) and GOOGLE_MAPS_API_KEY:
            url = ('https://maps.googleapis.com/maps/api/geocode/json'
                   '?address=' + quote_plus(row['address'] + ', Australia') +
                   '&key=' + GOOGLE_MAPS_API_KEY)
            geo = _http_get_json(url)
            if geo.get('status') == 'OK' and geo.get('results'):
                loc = geo['results'][0]['geometry']['location']
                lat, lng = loc['lat'], loc['lng']
                set_clauses += ['lat=?', 'lng=?']
                values += [lat, lng]
        if lat is not None and lng is not None:
            am = _compute_amenities(lat, lng)
            for col in ('hosp1_name','hosp1_km','hosp2_name','hosp2_km',
                        'hosp3_name','hosp3_km','medical_name','medical_km',
                        'shop_name','shop_km','train_name','train_km',
                        'bus_km','tram_name','tram_km'):
                set_clauses.append(col + '=?')
                values.append(am[col] if am[col] != '' else None)

    # Always stamp last_edited_*
    set_clauses += ['last_edited_by_id=?', 'last_edited_by_name=?', 'last_edited_at=?']
    values += [u['user_id'], u['full_name'], datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')]

    values.append(mid)
    db.execute('UPDATE manual_properties SET ' + ', '.join(set_clauses) + ' WHERE id=?', values)
    db.commit()

    updated = db.execute('SELECT * FROM manual_properties WHERE id=?', (mid,)).fetchone()
    db.close()
    log_event('manual_property_edited', {'id': mid, 'fields': list(cleaned.keys()),
                                          'recomputed': request.args.get('recompute_amenities','').lower()=='true'})
    return jsonify({'ok': True, 'property': _manual_row_to_dict(updated)})


@app.route('/api/manual-properties/<int:mid>', methods=['DELETE'])
@admin_required
def api_delete_manual_property(mid):
    """Permanently delete a manual property. Admin only.
    Also purges any matching hidden_properties row so the same address can
    be re-added later without a stale C1 hide blocking it."""
    db = get_db()
    row = db.execute(
        'SELECT id, state, address, address_key FROM manual_properties WHERE id=?', (mid,)
    ).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Manual property not found'}), 404
    state = row['state']
    addr_key = row['address_key']
    db.execute('DELETE FROM manual_properties WHERE id=?', (mid,))
    purged = db.execute(
        'DELETE FROM hidden_properties WHERE state=? AND address_key=?',
        (state, addr_key)
    ).rowcount
    db.commit()
    db.close()
    log_event('manual_property_deleted', {
        'id': mid, 'state': state, 'address': row['address'],
        'hide_rows_purged': purged,
    })
    return jsonify({'ok': True, 'hide_rows_purged': purged})


# ===========================================================================
# C3: SDA Supply intelligence (SA2-level)
# ---------------------------------------------------------------------------
# Two NDIS XLSX files (SDA-by-provider and Group-homes-by-provider) get
# uploaded together by an admin. Per SA2, we compute:
#   - Filled SDA places, vacancies, providers
#   - Filled group-home places, vacancies, providers (future SDA demand,
#     since group homes are being wound down and residents will need SDA)
#   - Net 'conversion deficit' = filled_gh - sda_vacancies
#   - Design-category mix of existing SDA stock
# Robust is excluded from BOTH sides because RPG doesn't build for Robust
# participants (they need dedicated specialty facilities, not HPS dwellings).
# The aggregated blob is stored once and served to every client on dashboard
# load — no per-row API calls.
# ===========================================================================

def _title_case_provider(name):
    """Some provider names come in ALL CAPS, others Title Case. Normalise to
    Title Case for display, but leave already-mixed names alone."""
    if not name:
        return ''
    s = str(name).strip()
    if any(c.islower() for c in s):
        return s
    out = []
    for w in s.split():
        if '&' in w or '/' in w:
            out.append(w)
        else:
            out.append(w.title())
    return ' '.join(out)


def _parse_supply_xlsx(file_storage, kind):
    """Parse one of the NDIS XLSX files into a list of dicts.
    `kind` is 'sda' or 'gh'; both files have the same column structure.
    Returns (rows, error_or_none, data_as_of_iso)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_storage, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return None, f'Could not open as XLSX: {e}', None

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return None, 'File is empty', None

    # Required columns
    required = ['State or Territory', 'SA4 Region', 'SA3 Region', 'SA2 Region',
                'Provider name', 'Design category', 'Places', 'Vacancies']
    missing = [c for c in required if c not in header]
    if missing:
        return None, f'Missing columns: {", ".join(missing)}', None

    col = {name: header.index(name) for name in header if name in required + ['Date', 'Dwelling type']}

    rows = []
    data_date = None
    for r in rows_iter:
        if not r or all(v is None for v in r):
            continue
        try:
            state = str(r[col['State or Territory']] or '').strip()
            sa4 = str(r[col['SA4 Region']] or '').strip()
            sa3 = str(r[col['SA3 Region']] or '').strip()
            sa2 = str(r[col['SA2 Region']] or '').strip()
            provider = str(r[col['Provider name']] or '').strip()
            design = str(r[col['Design category']] or '').strip()
            places = int(r[col['Places']] or 0)
            vacancies = int(r[col['Vacancies']] or 0)
            if 'Dwelling type' in col:
                dwelling = str(r[col['Dwelling type']] or '').strip()
            else:
                dwelling = ''
            if 'Date' in col and r[col['Date']]:
                d = r[col['Date']]
                if hasattr(d, 'isoformat'):
                    data_date = d.isoformat()[:10]
                else:
                    data_date = str(d)[:10]
        except (ValueError, IndexError, TypeError):
            continue
        if not state or not sa2:
            continue
        rows.append({
            'state': state, 'sa4': sa4, 'sa3': sa3, 'sa2': sa2,
            'provider': provider, 'design': design, 'dwelling': dwelling,
            'places': places, 'vacancies': vacancies,
        })

    return rows, None, data_date


def _aggregate_supply(sda_rows, gh_rows):
    """Combine per-row supply data into per-SA2 aggregates.
    Excludes Robust dwellings from both sides — RPG only builds HPS dwellings
    accepting HPS/FA/IL participants; Robust is a separate market segment.

    Returns {state_code: {sa2_name: aggregate_dict}}."""
    STATE_MAP = {
        'Victoria': 'vic',
        'New South Wales': 'nsw',
        'Queensland': 'qld',
    }

    # Filter Robust out of both inputs
    sda_rel = [r for r in sda_rows if r['design'] != 'Robust']
    gh_rel  = [r for r in gh_rows  if r['design'] != 'Robust']

    # Index by (state, sa2)
    by_sa2 = {}
    for r in sda_rel:
        st = STATE_MAP.get(r['state'])
        if not st:
            continue
        key = (st, r['sa2'])
        rec = by_sa2.setdefault(key, _empty_sa2_record(r['sa3'], r['sa4']))
        rec['sda_places'] += r['places']
        rec['sda_vacancies'] += r['vacancies']
        # Provider tally for SDA side
        prov = rec['sda_provider_idx'].setdefault(r['provider'], {'places': 0, 'vacancies': 0})
        prov['places'] += r['places']
        prov['vacancies'] += r['vacancies']
        # Design mix
        design_key = {
            'High physical support': 'HPS',
            'Fully accessible': 'FA',
            'Improved liveability': 'IL',
            'Multiple design category': 'Multi',
        }.get(r['design'])
        if design_key:
            rec['design_mix'][design_key] += r['places']

    for r in gh_rel:
        st = STATE_MAP.get(r['state'])
        if not st:
            continue
        key = (st, r['sa2'])
        rec = by_sa2.setdefault(key, _empty_sa2_record(r['sa3'], r['sa4']))
        rec['gh_places'] += r['places']
        rec['gh_vacancies'] += r['vacancies']
        prov = rec['gh_provider_idx'].setdefault(r['provider'], {'places': 0, 'vacancies': 0})
        prov['places'] += r['places']
        prov['vacancies'] += r['vacancies']

    # Finalise: compute derived fields and turn provider indices into sorted lists
    out = {'vic': {}, 'nsw': {}, 'qld': {}}
    for (st, sa2), rec in by_sa2.items():
        rec['sda_filled'] = rec['sda_places'] - rec['sda_vacancies']
        rec['gh_filled']  = rec['gh_places']  - rec['gh_vacancies']
        rec['deficit']    = rec['gh_filled'] - rec['sda_vacancies']
        # Convert provider dicts to lists sorted by places desc
        rec['sda_providers'] = sorted(
            ({'name': _title_case_provider(n), **v} for n, v in rec['sda_provider_idx'].items()),
            key=lambda x: x['places'], reverse=True,
        )
        rec['gh_providers'] = sorted(
            ({'name': _title_case_provider(n), **v} for n, v in rec['gh_provider_idx'].items()),
            key=lambda x: x['places'], reverse=True,
        )
        del rec['sda_provider_idx']
        del rec['gh_provider_idx']
        out[st][sa2] = rec
    return out


def _empty_sa2_record(sa3, sa4):
    return {
        'sa3': sa3, 'sa4': sa4,
        'sda_places': 0, 'sda_vacancies': 0, 'sda_filled': 0,
        'gh_places': 0, 'gh_vacancies': 0, 'gh_filled': 0,
        'deficit': 0,
        'sda_provider_idx': {},
        'gh_provider_idx': {},
        'design_mix': {'HPS': 0, 'FA': 0, 'IL': 0, 'Multi': 0},
    }


@app.route('/api/sda-supply')
@login_required
def api_sda_supply():
    """Return the full SA2 supply snapshot.

    Shape: {'as_of': '2025-12-31', 'data': {state: {sa2: {...aggregate}}}}.
    Small enough (~400KB JSON) to ship in one request — the client caches it
    in memory and joins it to property rows on render."""
    db = get_db()
    rows = db.execute(
        'SELECT state, sa2, data_json, data_as_of FROM sda_supply'
    ).fetchall()
    db.close()
    data = {'vic': {}, 'nsw': {}, 'qld': {}}
    as_of = None
    for r in rows:
        try:
            data[r['state']][r['sa2']] = json.loads(r['data_json'])
            if r['data_as_of'] and (not as_of or r['data_as_of'] > as_of):
                as_of = r['data_as_of']
        except (json.JSONDecodeError, KeyError):
            continue
    return jsonify({'as_of': as_of, 'data': data})


UPLOAD_SUPPLY_HTML = """<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Upload SDA Supply Data - SDA Screener</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Arial,sans-serif;background:#f3f4f6;color:#111}
.top{background:#002060;color:#fff;padding:14px 24px;display:flex;justify-content:space-between;align-items:center}
.top h1{font-size:15px;font-weight:700}
.top a{color:#93c5fd;font-size:13px;text-decoration:none;margin-left:16px}
.wrap{max-width:900px;margin:0 auto;padding:24px}
.card{background:#fff;border-radius:8px;padding:24px;box-shadow:0 1px 3px rgba(0,0,0,.1);margin-bottom:16px}
.card h2{font-size:15px;font-weight:700;color:#002060;margin-bottom:10px;padding-bottom:10px;border-bottom:2px solid #e5e7eb}
.help{font-size:12px;color:#374151;line-height:1.6;background:#f9fafb;padding:14px 18px;border-radius:6px;margin-bottom:18px;border-left:3px solid #0F6E56}
.help b{color:#0F6E56}
.field{margin-bottom:18px}
.field label{display:block;font-size:11px;color:#6B7280;font-weight:700;margin-bottom:6px;text-transform:uppercase;letter-spacing:.5px}
.field input[type=file]{width:100%;padding:10px;border:2px dashed #d1d5db;border-radius:8px;background:#f9fafb;font-size:12px;cursor:pointer}
.field input[type=file]:hover{border-color:#0F6E56;background:#f0fdf4}
.btn{padding:12px 24px;background:#0F6E56;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:700;width:100%}
.btn:hover{background:#0a5942}
.msg{padding:12px 16px;border-radius:6px;margin-bottom:16px;font-size:13px;line-height:1.5}
.ok{background:#dcfce7;color:#166534;border:1px solid #86efac}
.er{background:#fee2e2;color:#991b1b;border:1px solid #fca5a5}
.status-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:18px}
.status-card{border:2px solid #e5e7eb;border-radius:8px;padding:14px;text-align:center}
.status-card.loaded{border-color:#0F6E56;background:#f0fdf4}
.status-state{font-size:14px;font-weight:800;color:#002060;margin-bottom:6px}
.status-count{font-size:18px;color:#0F6E56;font-weight:700}
.status-empty{font-size:12px;color:#9CA3AF}
.status-date{font-size:10px;color:#6B7280;margin-top:6px}
.tabs{display:flex;gap:4px;margin-bottom:16px}
.tab{padding:10px 18px;border-radius:6px;font-size:13px;font-weight:600;background:#e5e7eb;color:#374151;text-decoration:none;display:inline-block}
.tab.on{background:#002060;color:#fff}
</style></head>
<body>
<div class="top">
  <h1>SDA Screener - Upload SDA Supply Data</h1>
  <div>
    <a href="/admin">Admin panel</a>
    <a href="/dashboard">Dashboard</a>
  </div>
</div>
<div class="wrap">
  <div class="tabs">
    <a href="/admin/upload_page" class="tab">Property data (CSV)</a>
    <a href="/admin/upload_sda_page" class="tab">SDA market</a>
    <a href="/admin/upload_supply_page" class="tab on">SDA supply (SA2)</a>
  </div>
  {{MSG_BLOCK}}
  <div class="card">
    <h2>Current data status</h2>
    <div class="status-grid">{{STATUS_GRID}}</div>
    {{AS_OF}}
  </div>
  <div class="card">
    <h2>Upload new snapshot</h2>
    <div class="help">
      Upload <b>both</b> NDIS quarterly XLSX files together. The server excludes Robust dwellings from both sides
      (since RPG only builds HPS dwellings) and computes the conversion deficit per SA2.
      Uploading replaces all existing supply data — old rows are cleared first.
      <br><br>
      <b>Source:</b> NDIS quarterly market position statements, "SA2 by provider" exports for SDA and Group Homes.
    </div>
    <form method="POST" action="/admin/upload_supply" enctype="multipart/form-data">
      <div class="field">
        <label for="sda_file">SDA file (sa2_by_provider_SDA.xlsx)</label>
        <input type="file" id="sda_file" name="sda_file" accept=".xlsx" required>
      </div>
      <div class="field">
        <label for="gh_file">Group homes file (sa2_by_provider_Group_Homes.xlsx)</label>
        <input type="file" id="gh_file" name="gh_file" accept=".xlsx" required>
      </div>
      <button class="btn" type="submit">Upload &amp; aggregate</button>
    </form>
  </div>
</div>
</body></html>"""


@app.route('/admin/upload_supply_page')
@admin_required
def upload_supply_page():
    """Render the SDA supply upload page with current data status."""
    db = get_db()
    rows = db.execute("""
        SELECT state, COUNT(*) AS sa2s, MAX(uploaded_at) AS uploaded_at,
               MAX(uploaded_by) AS uploaded_by, MAX(data_as_of) AS data_as_of
        FROM sda_supply GROUP BY state
    """).fetchall()
    db.close()
    status_by_state = {r['state']: dict(r) for r in rows}
    grid_html = ''
    for st, label in [('vic', 'VIC'), ('nsw', 'NSW'), ('qld', 'QLD')]:
        s = status_by_state.get(st)
        if s and s['sa2s']:
            grid_html += (f'<div class="status-card loaded">'
                          f'<div class="status-state">{label}</div>'
                          f'<div class="status-count">{s["sa2s"]} SA2s</div>'
                          f'<div class="status-date">uploaded {(s["uploaded_at"] or "")[:10]} by {s["uploaded_by"] or "?"}</div>'
                          f'</div>')
        else:
            grid_html += (f'<div class="status-card">'
                          f'<div class="status-state">{label}</div>'
                          f'<div class="status-empty">no data</div>'
                          f'</div>')
    overall_as_of = max(
        (r['data_as_of'] for r in rows if r['data_as_of']), default=None
    )
    as_of_html = (f'<div style="font-size:11px;color:#6B7280;text-align:center;margin-top:8px">'
                  f'Data as of {overall_as_of}</div>') if overall_as_of else ''

    msg = request.args.get('msg', '')
    msg_type = request.args.get('msg_type', 'ok')
    msg_block = f'<div class="msg {msg_type}">{msg}</div>' if msg else ''

    html = (UPLOAD_SUPPLY_HTML
            .replace('{{STATUS_GRID}}', grid_html)
            .replace('{{AS_OF}}', as_of_html)
            .replace('{{MSG_BLOCK}}', msg_block))
    return make_response(html)


@app.route('/admin/upload_supply', methods=['POST'])
@admin_required
def upload_supply():
    """Accept both NDIS XLSX files, aggregate, and replace the stored
    supply snapshot. This is an all-or-nothing replace so the dashboard
    always shows a consistent snapshot."""
    sda_file = request.files.get('sda_file')
    gh_file = request.files.get('gh_file')
    if not sda_file or not gh_file:
        return redirect('/admin/upload_supply_page?msg=Both+files+required&msg_type=er')

    sda_rows, err, sda_date = _parse_supply_xlsx(sda_file, 'sda')
    if err:
        return redirect(f'/admin/upload_supply_page?msg=SDA+file:+{quote_plus(err)}&msg_type=er')
    gh_rows, err, gh_date = _parse_supply_xlsx(gh_file, 'gh')
    if err:
        return redirect(f'/admin/upload_supply_page?msg=Group+home+file:+{quote_plus(err)}&msg_type=er')

    # Sanity check: file looks like NDIS data?
    if len(sda_rows) < 100:
        return redirect(f'/admin/upload_supply_page?msg=SDA+file+only+had+{len(sda_rows)}+rows+(expected+thousands)&msg_type=er')
    if len(gh_rows) < 50:
        return redirect(f'/admin/upload_supply_page?msg=Group+home+file+only+had+{len(gh_rows)}+rows&msg_type=er')

    aggregated = _aggregate_supply(sda_rows, gh_rows)
    data_as_of = sda_date or gh_date

    u = request.current_user
    by_user = u.get('full_name') or u.get('username') or '?'

    db = get_db()
    db.execute('DELETE FROM sda_supply')
    total_inserted = 0
    for st, sa2_map in aggregated.items():
        for sa2, rec in sa2_map.items():
            db.execute(
                """INSERT INTO sda_supply
                   (state, sa2, sa3, sa4, data_json, data_as_of, uploaded_by)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (st, sa2, rec.get('sa3', ''), rec.get('sa4', ''),
                 json.dumps(rec, separators=(',', ':')),
                 data_as_of, by_user)
            )
            total_inserted += 1
    db.commit()
    db.close()
    log_event('sda_supply_uploaded', {
        'sa2_count': total_inserted, 'data_as_of': data_as_of,
        'sda_rows': len(sda_rows), 'gh_rows': len(gh_rows),
    })
    return redirect(f'/admin/upload_supply_page?msg=Loaded+{total_inserted}+SA2s+(data+as+of+{data_as_of or "unknown"})&msg_type=ok')


@app.route('/admin/upload_dashboard', methods=['POST'])
@admin_required
def upload_dashboard():
    f = request.files.get('htmlfile')
    if not f: return redirect('/admin/upload_page?msg=No+file&msg_type=er')
    html = f.read().decode('utf-8')
    db = get_db()
    db.execute("INSERT INTO config (key,value) VALUES ('dashboard_html',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=datetime('now')", (html,))
    db.commit()
    db.close()
    return redirect('/admin/upload_page?msg=Dashboard+updated+successfully&msg_type=ok')

@app.route('/admin')
@app.route('/admin/<section>')
@admin_required
def admin(section='activity'):
    db = get_db()
    total_users = db.execute("SELECT COUNT(*) FROM users WHERE role='team'").fetchone()[0]
    total_sessions = db.execute("SELECT COUNT(*) FROM sessions").fetchone()[0]
    views = db.execute("SELECT COUNT(*) FROM events WHERE event_type='property_view'").fetchone()[0]
    dls = db.execute("SELECT COUNT(*) FROM events WHERE event_type IN ('csv_download','docx_download')").fetchone()[0]
    stats = {'total_users':total_users,'total_sessions':total_sessions,'views':views,'downloads':dls}
    users = [dict(r) for r in db.execute("""
        SELECT u.id,u.full_name,u.username,u.role,u.active,
               COUNT(DISTINCT s.id) as session_count, MAX(s.login_at) as last_login,
               SUM(CASE WHEN s.logout_at IS NOT NULL
                   THEN ROUND((julianday(s.logout_at)-julianday(s.login_at))*24*60)
                   ELSE NULL END) as total_minutes
        FROM users u LEFT JOIN sessions s ON s.user_id=u.id
        GROUP BY u.id ORDER BY last_login DESC
    """).fetchall()]
    all_users = [dict(r) for r in db.execute('SELECT * FROM users ORDER BY full_name').fetchall()]
    prop_views = []
    for e in db.execute("SELECT e.*,u.full_name FROM events e JOIN users u ON e.user_id=u.id WHERE e.event_type='property_view' ORDER BY e.created_at DESC LIMIT 200").fetchall():
        row = dict(e)
        try:
            d = json.loads(row.get('event_data') or '{}')
            row['addr']=d.get('addr','n/a'); row['state']=d.get('state','').upper()
            row['sa4']=d.get('sa4','n/a'); row['price']='$'+str(int(d['price'])) if d.get('price') else 'n/a'
        except: row['addr']=row['state']=row['sa4']=row['price']='n/a'
        prop_views.append(row)
    dl_list = []
    for e in db.execute("SELECT e.*,u.full_name FROM events e JOIN users u ON e.user_id=u.id WHERE e.event_type IN ('csv_download','docx_download') ORDER BY e.created_at DESC LIMIT 200").fetchall():
        row = dict(e)
        try:
            d = json.loads(row.get('event_data') or '{}')
            row['addr']=d.get('addr','n/a'); row['state']=d.get('state','').upper()
        except: row['addr']=row['state']='n/a'
        dl_list.append(row)
    # Hidden properties — always fetch the count (for the tab badge), and the
    # full list only if we're rendering that section.
    hidden_count = db.execute('SELECT COUNT(*) FROM hidden_properties').fetchone()[0]
    hidden = []
    if section == 'hidden':
        hidden = [dict(r) for r in db.execute(
            """SELECT id, state, address, hidden_by_id, hidden_by_name,
                      hidden_at, reason
               FROM hidden_properties ORDER BY hidden_at DESC"""
        ).fetchall()]
    db.close()
    msg = request.args.get('msg','')
    msg_type = request.args.get('msg_type','ok')
    # Check if dashboard is loaded
    dash_loaded = bool(get_dashboard_html() != '<h1>Dashboard not loaded. Upload via admin panel.</h1>')
    from jinja2 import Template as T
    return make_response(T(ADMIN_HTML).render(
        current_user=request.current_user, section=section,
        stats=stats, users=users, all_users=all_users,
        prop_views=prop_views, downloads=dl_list,
        hidden=hidden, hidden_count=hidden_count,
        msg=msg, msg_type=msg_type, dash_loaded=dash_loaded
    ))

@app.route('/admin/upload_page')
@admin_required
def upload_page():
    db = get_db()
    rows = db.execute('SELECT state,row_count,uploaded_at,uploaded_by FROM property_data').fetchall()
    db.close()
    status = {r['state']: dict(r) for r in rows}
    msg = request.args.get('msg','')
    msg_type = request.args.get('msg_type','ok')
    html = UPLOAD_HTML.replace('{{STATUS}}', json.dumps(status)).replace('{{MSG}}', msg).replace('{{MSG_TYPE}}', msg_type)
    return make_response(html)

@app.route('/admin/upload', methods=['POST'])
@admin_required
def upload_csv():
    state = request.form.get('state','').lower()
    if state not in ('vic','nsw','qld'): return redirect('/admin/upload_page?msg=Invalid+state&msg_type=er')
    f = request.files.get('csvfile')
    if not f: return redirect('/admin/upload_page?msg=No+file+selected&msg_type=er')
    try:
        csv_data = f.read().decode('utf-8')
        rows = len([l for l in csv_data.strip().split('\n') if l]) - 1
        db = get_db()
        db.execute("""INSERT INTO property_data (state,csv_data,row_count,uploaded_by) VALUES (?,?,?,?)
                      ON CONFLICT(state) DO UPDATE SET csv_data=excluded.csv_data,
                      row_count=excluded.row_count,uploaded_by=excluded.uploaded_by,
                      uploaded_at=datetime('now')""",
                   (state, csv_data, rows, request.current_user['username']))
        db.commit()
        db.close()
        return redirect('/admin/upload_page?msg='+state.upper()+'+uploaded+('+str(0)+'+properties)&msg_type=ok')
    except: return redirect('/admin/upload_page?msg=Upload+failed&msg_type=er')


# =============================================================================
# SDA Market Map — three layers (Radius / Existing / Permitted)
# Stores per-property pin data for the SDA market map. CSVs come from
# LandChecker exports which already have lat/lng in a 'Coordinates' column,
# so no geocoding is needed at upload time.
# =============================================================================

def _parse_landchecker_csv(csv_text):
    """Parse a LandChecker-style CSV. Returns list of dicts ready for insertion.
    Required column: Address. Coordinates column is "lng, lat" pair string."""
    import csv as _csv, io
    reader = _csv.DictReader(io.StringIO(csv_text))
    rows = []
    skipped = 0
    for raw in reader:
        # Normalise keys (case + whitespace insensitive)
        r = {(k or '').strip().lower(): (v or '').strip() for k, v in raw.items()}
        addr = r.get('address') or ''
        if not addr:
            skipped += 1
            continue
        # Coordinates are "lng, lat" e.g. "145.17783486692818, -37.976366172276954"
        lat, lng = None, None
        coords = r.get('coordinates') or ''
        if coords:
            parts = [p.strip() for p in coords.split(',')]
            if len(parts) == 2:
                try:
                    lng = float(parts[0])
                    lat = float(parts[1])
                except ValueError:
                    pass
        # Combine 3 Note fields into one searchable notes blob
        note_blob = []
        for n in ('1', '2', '3'):
            txt = r.get(f'note {n}') or ''
            author = r.get(f'note {n} author') or ''
            date = r.get(f'note {n} date') or ''
            if txt:
                if author or date:
                    note_blob.append(f'[{date} {author}] {txt}'.strip())
                else:
                    note_blob.append(txt)
        try:
            area = float(r.get('area') or 0) or None
        except ValueError:
            area = None
        try:
            frontage = float(r.get('frontage (m)') or r.get('frontage') or 0) or None
        except ValueError:
            frontage = None
        rows.append({
            'address': addr,
            'suburb': (r.get('suburb') or '').upper(),
            'state': (r.get('state') or '').lower(),
            'postcode': r.get('postcode') or '',
            'area_m2': area,
            'frontage_m': frontage,
            'planning_zones': r.get('planning zones') or '',
            'notes': '\n'.join(note_blob) if note_blob else '',
            'lat': lat,
            'lng': lng,
        })
    return rows, skipped


def _parse_radius_pipeline_csv(csv_text):
    """Parse the Radius internal pipeline CSV format (Name / Dwelling / Location).

    Format has nested sections with section headers, the column header row
    'Name,Dwelling,Location', and 'New Item' separators. We extract any row that
    has a non-empty Location, treating it as a property regardless of section.
    Returns list of dicts (without lat/lng — caller must geocode)."""
    import csv as _csv, io
    reader = _csv.reader(io.StringIO(csv_text))
    rows = []
    skipped = 0
    section = None
    for raw in reader:
        cells = [(c or '').strip() for c in raw]
        if not cells or all(c == '' for c in cells):
            continue
        # Section header (single non-empty cell, rest empty)
        if cells[0] and (len(cells) == 1 or not any(cells[1:])):
            section = cells[0]
            continue
        # Column header row — skip
        if cells[0].lower() == 'name' and len(cells) > 1 and cells[1].lower() == 'dwelling':
            continue
        # Padding row 'New Item,,'
        if cells[0].lower() == 'new item':
            continue
        # Data row: Name, Dwelling, Location
        name = cells[0]
        dwelling = cells[1] if len(cells) > 1 else ''
        location = cells[2] if len(cells) > 2 else ''
        # Need at least a location to geocode
        if not location:
            skipped += 1
            continue
        # Try to parse a state code from the location string
        state = ''
        for st in ('VIC', 'NSW', 'QLD'):
            if st in location.upper():
                state = st.lower()
                break
        # Notes: Dwelling type + section context
        notes_parts = []
        if dwelling:
            notes_parts.append('Dwelling: ' + dwelling)
        if section and section not in ('SDA Project Pipeline', 'All SDA'):
            notes_parts.append('Status: ' + section)
        rows.append({
            'address': location,             # use the cleaner Google-formatted location for geocoding
            'suburb': '',                    # filled in if geocoder returns it
            'state': state,
            'postcode': '',
            'area_m2': None,
            'frontage_m': None,
            'planning_zones': '',
            'notes': '\n'.join(notes_parts),
            'lat': None,
            'lng': None,
        })
    return rows, skipped


def _detect_csv_format(csv_text):
    """Return 'landchecker' or 'radius_pipeline' based on header inspection."""
    head = csv_text[:2000].lower()
    if 'coordinates' in head and 'planning zones' in head:
        return 'landchecker'
    if 'sda project pipeline' in head or ('name,dwelling,location' in head):
        return 'radius_pipeline'
    # Fallback: assume landchecker if it has 'address' as the first header
    return 'landchecker'


def _geocode_address(addr):
    """Server-side geocode a single address. Returns (lat, lng) or (None, None)."""
    if not addr or not GOOGLE_MAPS_API_KEY:
        return (None, None)
    try:
        url = ('https://maps.googleapis.com/maps/api/geocode/json'
               '?address=' + quote_plus(addr) + '&region=au'
               '&key=' + GOOGLE_MAPS_API_KEY)
        data = _http_get_json(url, timeout=10)
        if data.get('status') == 'OK' and data.get('results'):
            loc = data['results'][0].get('geometry', {}).get('location', {})
            if 'lat' in loc and 'lng' in loc:
                return (loc['lat'], loc['lng'])
    except Exception:
        pass
    return (None, None)


@app.route('/admin/upload_sda', methods=['POST'])
@admin_required
def upload_sda_market():
    """Upload an SDA market CSV. Accepts either:
      - LandChecker export format (has Coordinates column — no geocoding needed)
      - Radius pipeline format (Name / Dwelling / Location — geocoded server-side)
    Form fields: layer (radius|existing|permitted), state (vic|nsw|qld), csvfile"""
    layer = (request.form.get('layer') or '').lower()
    state = (request.form.get('state') or '').lower()
    if layer not in ('radius', 'existing', 'ghomes', 'permitted'):
        return redirect('/admin/upload_sda_page?msg=Invalid+layer&msg_type=er')
    if state not in ('vic', 'nsw', 'qld'):
        return redirect('/admin/upload_sda_page?msg=Invalid+state&msg_type=er')
    f = request.files.get('csvfile')
    if not f:
        return redirect('/admin/upload_sda_page?msg=No+file+selected&msg_type=er')
    try:
        csv_text = f.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        try:
            csv_text = f.read().decode('latin-1')
        except Exception:
            return redirect('/admin/upload_sda_page?msg=Could+not+decode+file&msg_type=er')

    fmt = _detect_csv_format(csv_text)
    try:
        if fmt == 'radius_pipeline':
            rows, skipped = _parse_radius_pipeline_csv(csv_text)
        else:
            rows, skipped = _parse_landchecker_csv(csv_text)
    except Exception as e:
        return redirect(f'/admin/upload_sda_page?msg=Parse+error:+{e}&msg_type=er')
    if not rows:
        return redirect('/admin/upload_sda_page?msg=No+rows+parsed+(check+CSV+structure)&msg_type=er')

    # Geocode any rows that don't already have lat/lng (Radius pipeline format)
    geocode_failed = 0
    for r in rows:
        if r['lat'] is None or r['lng'] is None:
            lat, lng = _geocode_address(r['address'])
            if lat is not None:
                r['lat'] = lat
                r['lng'] = lng
            else:
                geocode_failed += 1

    db = get_db()
    # For Radius pipeline format, the CSV may span multiple states. We delete
    # all rows for this layer (across all states) and let the per-row state
    # detection determine where each pin ends up. For LandChecker format,
    # only delete the specific (layer, state) combo.
    if fmt == 'radius_pipeline':
        db.execute('DELETE FROM sda_market WHERE layer=?', (layer,))
    else:
        db.execute('DELETE FROM sda_market WHERE layer=? AND state=?', (layer, state))
    by = request.current_user['username']
    inserted = 0
    state_counts = {}
    for r in rows:
        # For Radius pipeline format, prefer the auto-detected state from each row.
        # For LandChecker, force form-selected state (handles mixed CSVs gracefully).
        row_state = r['state'] if (fmt == 'radius_pipeline' and r['state']) else state
        state_counts[row_state] = state_counts.get(row_state, 0) + 1
        db.execute("""INSERT INTO sda_market
                      (layer, state, address, suburb, postcode, area_m2, frontage_m,
                       planning_zones, notes, lat, lng, uploaded_by)
                      VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                   (layer, row_state, r['address'], r['suburb'], r['postcode'],
                    r['area_m2'], r['frontage_m'], r['planning_zones'],
                    r['notes'], r['lat'], r['lng'], by))
        inserted += 1
    db.commit()
    db.close()
    msg_parts = [f'{layer.upper()}+uploaded:+{inserted}+pins']
    if state_counts and fmt == 'radius_pipeline':
        bd = '+'.join(f'{k.upper()}:{v}' for k, v in sorted(state_counts.items()))
        msg_parts.append('split:+' + bd)
    if skipped:
        msg_parts.append(f'{skipped}+skipped')
    if geocode_failed:
        msg_parts.append(f'{geocode_failed}+failed+to+geocode')
    if fmt == 'radius_pipeline':
        msg_parts.append('(Radius+pipeline+format,+geocoded+server-side)')
    return redirect('/admin/upload_sda_page?msg=' + '+•+'.join(msg_parts) + '&msg_type=ok')


@app.route('/api/sda-market')
@login_required
def api_sda_market():
    """Return all SDA market pins.
    Query params (optional): layer, state — to filter."""
    layer = (request.args.get('layer') or '').lower()
    state = (request.args.get('state') or '').lower()
    sql = ('SELECT id, layer, state, address, suburb, postcode, area_m2, frontage_m, '
           'planning_zones, notes, lat, lng FROM sda_market WHERE lat IS NOT NULL AND lng IS NOT NULL')
    args = []
    if layer in ('radius', 'existing', 'ghomes', 'permitted'):
        sql += ' AND layer=?'
        args.append(layer)
    if state in ('vic', 'nsw', 'qld'):
        sql += ' AND state=?'
        args.append(state)
    db = get_db()
    rows = db.execute(sql, args).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/sda-market/status')
@login_required
def api_sda_market_status():
    """Return per-(layer,state) pin counts and last-upload metadata."""
    db = get_db()
    rows = db.execute("""
        SELECT layer, state, COUNT(*) AS pins, MAX(uploaded_at) AS uploaded_at,
               MAX(uploaded_by) AS uploaded_by
        FROM sda_market GROUP BY layer, state
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/sda-nearby')
@login_required
def api_sda_nearby():
    """Find SDA market pins within radius_km of (lat, lng).
    Query: lat, lng, radius_km (default 5)
    Returns: { 'radius':[...], 'existing':[...], 'permitted':[...] } where
    each item has address, suburb, state, area_m2, frontage_m, planning_zones,
    notes, lat, lng, distance_km. Sorted by distance ascending."""
    try:
        lat = float(request.args.get('lat', ''))
        lng = float(request.args.get('lng', ''))
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid lat/lng'}), 400
    try:
        radius_km = float(request.args.get('radius_km') or 5.0)
    except ValueError:
        radius_km = 5.0
    radius_km = max(0.1, min(50.0, radius_km))

    # Pre-filter with a bounding box to avoid running haversine on every row.
    # 1 deg latitude  ~= 111 km
    # 1 deg longitude ~= 111 km * cos(lat)
    import math
    dlat = radius_km / 111.0
    dlng = radius_km / (111.0 * max(0.001, math.cos(math.radians(lat))))
    db = get_db()
    rows = db.execute("""
        SELECT layer, state, address, suburb, postcode, area_m2, frontage_m,
               planning_zones, notes, lat, lng
        FROM sda_market
        WHERE lat IS NOT NULL AND lng IS NOT NULL
          AND lat BETWEEN ? AND ?
          AND lng BETWEEN ? AND ?
    """, (lat - dlat, lat + dlat, lng - dlng, lng + dlng)).fetchall()
    db.close()

    out = {'radius': [], 'existing': [], 'ghomes': [], 'permitted': []}
    for r in rows:
        d = _haversine_km(lat, lng, r['lat'], r['lng'])
        if d > radius_km:
            continue
        item = dict(r)
        item['distance_km'] = round(d, 2)
        layer = item['layer']
        if layer in out:
            out[layer].append(item)
    # Sort each layer by distance
    for k in out:
        out[k].sort(key=lambda x: x['distance_km'])
    return jsonify(out)


@app.route('/admin/upload_sda_page')
@admin_required
def upload_sda_page():
    """Render the SDA market upload page."""
    db = get_db()
    rows = db.execute("""
        SELECT layer, state, COUNT(*) AS pins, MAX(uploaded_at) AS uploaded_at,
               MAX(uploaded_by) AS uploaded_by
        FROM sda_market GROUP BY layer, state
    """).fetchall()
    db.close()
    status = {}
    for r in rows:
        status.setdefault(r['layer'], {})[r['state']] = {
            'pins': r['pins'],
            'uploaded_at': r['uploaded_at'],
            'uploaded_by': r['uploaded_by']
        }
    msg = request.args.get('msg', '')
    msg_type = request.args.get('msg_type', 'ok')
    html = (UPLOAD_SDA_HTML
            .replace('{{STATUS}}', json.dumps(status))
            .replace('{{MSG}}', msg)
            .replace('{{MSG_TYPE}}', msg_type))
    return make_response(html)


@app.route('/admin/users/add_form', methods=['POST'])
@admin_required
def add_user_form():
    full_name = request.form.get('full_name','').strip()
    username = request.form.get('username','').strip().lower()
    password = request.form.get('password','')
    role = request.form.get('role','team')
    if not full_name or not username or len(password) < 6:
        return redirect('/admin/users?msg=All+fields+required&msg_type=er')
    try:
        db = get_db()
        db.execute('INSERT INTO users (username,password_hash,full_name,role) VALUES (?,?,?,?)',
                   (username, generate_password_hash(password), full_name, role))
        db.commit(); db.close()
        return redirect('/admin/users?msg='+full_name+'+added&msg_type=ok')
    except sqlite3.IntegrityError:
        return redirect('/admin/users?msg=Username+exists&msg_type=er')

@app.route('/admin/users/add', methods=['POST'])
@admin_required
def add_user():
    data = request.get_json()
    try:
        db = get_db()
        db.execute('INSERT INTO users (username,password_hash,full_name,role) VALUES (?,?,?,?)',
                   (data['username'].lower(), generate_password_hash(data['password']), data['full_name'], data.get('role','team')))
        db.commit(); db.close()
        return jsonify({'ok': True})
    except: return jsonify({'error': 'Failed'}), 400

@app.route('/admin/users/<int:uid>/password', methods=['POST'])
@admin_required
def change_password(uid):
    pw = (request.get_json(silent=True) or {}).get('password') or request.form.get('password','')
    if len(pw) < 6:
        return redirect('/admin/users?msg=Password+too+short&msg_type=er') if 'json' not in (request.content_type or '') else (jsonify({'error':'Too short'}), 400)
    db = get_db()
    db.execute('UPDATE users SET password_hash=? WHERE id=?', (generate_password_hash(pw), uid))
    db.commit(); db.close()
    if 'json' in (request.content_type or ''): return jsonify({'ok': True})
    return redirect('/admin/users?msg=Password+updated&msg_type=ok')

@app.route('/admin/users/<int:uid>/toggle', methods=['POST'])
@admin_required
def toggle_user(uid):
    db = get_db()
    user = db.execute('SELECT active FROM users WHERE id=?', (uid,)).fetchone()
    if user:
        db.execute('UPDATE users SET active=? WHERE id=?', (1-user['active'], uid))
        db.commit()
    db.close()
    if 'json' in (request.content_type or ''): return jsonify({'ok': True})
    return redirect('/admin/users?msg=Updated&msg_type=ok')

@app.route('/admin/users/<int:uid>/toggle_role', methods=['POST'])
@admin_required
def toggle_role(uid):
    # Flip a user's role between 'admin' and 'team'. Two safety guards:
    #   1. Admins can't demote themselves (would lock them out of admin pages
    #      immediately on next click).
    #   2. The last remaining admin can't be demoted at all, preventing the
    #      system from ending up with zero admins.
    me = getattr(request, 'current_user', None) or {}
    if me.get('user_id') == uid:
        return redirect('/admin/users?msg=You+cannot+change+your+own+role&msg_type=er')
    db = get_db()
    user = db.execute('SELECT role,full_name FROM users WHERE id=?', (uid,)).fetchone()
    if not user:
        db.close()
        return redirect('/admin/users?msg=User+not+found&msg_type=er')
    new_role = 'team' if user['role'] == 'admin' else 'admin'
    if new_role == 'team':
        # Guard against demoting the last active admin.
        admin_count = db.execute("SELECT COUNT(*) AS c FROM users WHERE role='admin' AND active=1").fetchone()['c']
        if admin_count <= 1:
            db.close()
            return redirect('/admin/users?msg=Cannot+demote+the+last+admin&msg_type=er')
    db.execute('UPDATE users SET role=? WHERE id=?', (new_role, uid))
    db.commit(); db.close()
    msg = user['full_name'].replace(' ', '+') + '+is+now+' + new_role
    return redirect('/admin/users?msg=' + msg + '&msg_type=ok')

@app.route('/admin/api/stats')
@admin_required
def admin_stats():
    db = get_db()
    users = db.execute("""
        SELECT u.id,u.full_name,u.username,u.role,u.active,
               COUNT(DISTINCT s.id) as session_count,MAX(s.login_at) as last_login
        FROM users u LEFT JOIN sessions s ON s.user_id=u.id GROUP BY u.id
    """).fetchall()
    events = db.execute("SELECT e.*,u.full_name,u.username FROM events e JOIN users u ON e.user_id=u.id ORDER BY e.created_at DESC LIMIT 200").fetchall()
    db.close()
    return jsonify({'stats':{'total_users':0,'total_sessions':0,'total_events':0},
                    'users':[dict(u) for u in users],'events':[dict(e) for e in events]})

@app.route('/admin/api/users')
@admin_required
def admin_users_api():
    db = get_db()
    users = db.execute('SELECT id,username,full_name,role,active FROM users ORDER BY full_name').fetchall()
    db.close()
    return jsonify([dict(u) for u in users])

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
